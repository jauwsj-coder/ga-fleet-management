from __future__ import annotations

import io
import logging
import json
import os
import shutil
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:  # PostgreSQL support is only required when DATABASE_URL is configured.
    psycopg2 = None
    RealDictCursor = None


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "ga_operations.db"
RAW_DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
DATABASE_URL = RAW_DATABASE_URL
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
elif DATABASE_URL.startswith("postgresql+psycopg2://"):
    DATABASE_URL = DATABASE_URL.replace("postgresql+psycopg2://", "postgresql://", 1)
USE_POSTGRES = DATABASE_URL.startswith("postgresql://")
REVIEW_PROOF_DIR = BASE_DIR / "uploads" / "review_follow_up"
P2H_ATTACHMENT_DIR = BASE_DIR / "uploads" / "p2h"
BACKUP_DIR = BASE_DIR / "backups"
APP_TABLES = [
    "employees",
    "employee_roles",
    "drivers",
    "vehicles",
    "vehicle_maintenance_history",
    "trip_requests",
    "option_lists",
    "app_meta",
    "trip_edit_logs",
    "p2h_reports",
    "p2h_checklist_items",
    "p2h_attachments",
    "p2h_holidays",
    "p2h_workday_overrides",
]

app = Flask(__name__)
app.secret_key = "ga-operations-local-dev"

logging.basicConfig(level=logging.DEBUG, format="%(levelname)s:%(name)s:%(message)s")
logger = logging.getLogger(__name__)


STATUS_PENDING = "pending_leader_approval"
STATUS_REJECTED = "rejected"
STATUS_APPROVED = "approved"
STATUS_PROCESSING = "processing_ga"
STATUS_ASSIGNED = "assigned"
STATUS_ON_TRIP = "on_trip"
STATUS_COMPLETED = "completed"
STATUS_REVIEWED = "reviewed"
STATUS_CANCELED = "canceled"
EDITABLE_BOOKING_STATUSES = {STATUS_PENDING, STATUS_REJECTED}
ACTIVE_BOOKING_STATUSES = {STATUS_PENDING, STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP}

STATUS_ALIASES = {
    "pending": STATUS_PENDING,
    "pending leader approval": STATUS_PENDING,
    "pending_leader_approval": STATUS_PENDING,
    "PENDING LEADER APPROVAL": STATUS_PENDING,
    "Pending Leader Approval": STATUS_PENDING,
    "approved": STATUS_APPROVED,
    "approved by leader": STATUS_APPROVED,
    "APPROVED BY LEADER": STATUS_APPROVED,
    "Approved": STATUS_APPROVED,
    "rejected": STATUS_REJECTED,
    "REJECTED": STATUS_REJECTED,
    "processing ga": STATUS_PROCESSING,
    "processing_ga": STATUS_PROCESSING,
    "PROCESSING GA": STATUS_PROCESSING,
    "assigned": STATUS_ASSIGNED,
    "assigned to driver": STATUS_ASSIGNED,
    "ASSIGNED TO DRIVER": STATUS_ASSIGNED,
    "on trip": STATUS_ON_TRIP,
    "on_trip": STATUS_ON_TRIP,
    "ON TRIP": STATUS_ON_TRIP,
    "completed": STATUS_COMPLETED,
    "COMPLETED": STATUS_COMPLETED,
    "reviewed": STATUS_REVIEWED,
    "canceled": STATUS_CANCELED,
    "cancelled": STATUS_CANCELED,
}

ROLE_LABELS = {
    "user": "User",
    "pimpinan": "Pimpinan",
    "ga_admin": "GA Admin",
    "driver": "Driver",
    "super_admin": "Super Admin GA",
}

MAINTENANCE_TYPE_OPTIONS = [
    "Service berkala",
    "Ganti oli",
    "Ganti filter",
    "Tune up",
    "Perbaikan rem",
    "Perbaikan AC",
    "Ganti ban",
    "Kelistrikan",
    "Spooring/balancing",
    "Body repair",
    "Lainnya",
]

MAINTENANCE_PART_OPTIONS = [
    "Oli mesin",
    "Filter oli",
    "Filter udara",
    "Filter AC",
    "Busi",
    "Kampas rem",
    "Minyak rem",
    "Ban",
    "Aki",
    "Wiper",
    "Lampu",
    "Radiator/coolant",
    "Shockbreaker",
    "Belt",
    "Kopling",
    "AC",
    "Kelistrikan",
    "Spooring/balancing",
    "Lainnya",
]

P2H_CHECKLIST = {
    "Pemeriksaan Mesin & Cairan": [
        "Oli mesin",
        "Air radiator/coolant",
        "Minyak rem",
        "Minyak power steering",
        "Air wiper",
        "Kondisi aki/battery",
        "Tidak ada kebocoran oli/cairan",
    ],
    "Pemeriksaan Ban & Kaki-kaki": [
        "Tekanan ban",
        "Kondisi tapak ban",
        "Ban cadangan",
        "Baut roda",
        "Suspensi/shockbreaker",
        "Tidak ada bunyi abnormal pada kaki-kaki",
    ],
    "Pemeriksaan Rem & Kemudi": [
        "Rem utama",
        "Rem tangan/parking brake",
        "Setir/kemudi normal",
        "Pedal rem normal",
        "Tidak ada getaran abnormal saat pengereman",
    ],
    "Pemeriksaan Lampu & Kelistrikan": [
        "Lampu utama",
        "Lampu sein",
        "Lampu rem",
        "Lampu mundur",
        "Klakson",
        "Wiper",
        "Panel indikator dashboard",
        "AC kendaraan",
    ],
    "Pemeriksaan Safety Equipment": [
        "APAR",
        "Kotak P3K",
        "Segitiga pengaman",
        "Dongkrak",
        "Kunci roda",
        "Sabuk pengaman",
        "Dokumen kendaraan/STNK",
    ],
    "Pemeriksaan Body & Interior": [
        "Body kendaraan",
        "Spion",
        "Kaca depan/samping/belakang",
        "Pintu dan central lock",
        "Jok/kursi",
        "Dashboard",
        "Bagasi",
    ],
    "Kebersihan Kendaraan": [
        "Eksterior bersih",
        "Interior bersih",
        "Karpet bersih",
        "Tidak ada sampah",
        "Tidak ada bau tidak sedap",
        "Bagasi bersih",
    ],
    "Kesiapan Operasional": [
        "BBM cukup",
        "KM awal diinput",
        "Kendaraan siap jalan",
        "Tidak ada suara/indikasi abnormal",
        "Driver menyatakan kendaraan layak digunakan",
    ],
}

P2H_RESULTS = {"OK", "Tidak OK", "Tidak Berlaku"}
P2H_STATUS_NORMAL = "Normal"
P2H_STATUS_FOLLOW_UP = "Perlu Follow Up GA"
P2H_FOLLOW_NEW = "Baru"
P2H_FOLLOW_PROCESS = "Diproses"
P2H_FOLLOW_DONE = "Selesai"
P2H_FOLLOW_REJECTED = "Ditolak / Tidak Valid"
DEFAULT_P2H_HOLIDAYS = {
    # Hari libur nasional dan cuti bersama 2026 berdasarkan SKB 3 Menteri yang dipublikasikan Kemensetneg.
    "2026-01-01": "Tahun Baru 2026",
    "2026-01-16": "Isra Mikraj Nabi Muhammad saw.",
    "2026-02-16": "Cuti Bersama Tahun Baru Imlek",
    "2026-02-17": "Tahun Baru Imlek",
    "2026-03-18": "Cuti Bersama Nyepi",
    "2026-03-19": "Hari Suci Nyepi",
    "2026-03-20": "Cuti Bersama Idulfitri",
    "2026-03-21": "Idulfitri",
    "2026-03-22": "Idulfitri",
    "2026-03-23": "Cuti Bersama Idulfitri",
    "2026-03-24": "Cuti Bersama Idulfitri",
    "2026-04-03": "Wafat Yesus Kristus",
    "2026-04-05": "Paskah",
    "2026-05-01": "Hari Buruh Internasional",
    "2026-05-14": "Kenaikan Yesus Kristus",
    "2026-05-15": "Cuti Bersama Kenaikan Yesus Kristus",
    "2026-05-27": "Iduladha",
    "2026-05-28": "Cuti Bersama Iduladha",
    "2026-05-31": "Hari Raya Waisak",
    "2026-06-01": "Hari Lahir Pancasila",
    "2026-06-16": "Tahun Baru Islam",
    "2026-08-17": "Proklamasi Kemerdekaan",
    "2026-08-25": "Maulid Nabi Muhammad saw.",
    "2026-12-24": "Cuti Bersama Natal",
    "2026-12-25": "Natal",
}

FULL_BOOKED_ALERT = """Terima kasih atas pengajuan pemesanan kendaraan.

Saat ini kendaraan dan/atau driver pada jadwal yang dipilih sedang tidak tersedia.

Apabila perjalanan bersifat mendesak, silakan menghubungi tim GA untuk pengecekan dan penyesuaian jadwal secara manual. Penyesuaian dapat dilakukan dalam estimasi maksimal 2 jam sebelum atau sesudah jadwal yang telah terdaftar, namun ketersediaan kendaraan dan driver tetap tidak dapat dipastikan.

Untuk informasi detail jadwal dan ketersediaan driver, silakan melihat Dashboard Jadwal Driver."""

FULL_BOOKED_ALERT_EN = """Thank you for submitting a vehicle booking request.

The vehicle and/or driver is currently unavailable for the selected schedule.

If the trip is urgent, please contact the GA team for manual schedule checking and adjustment. Adjustments may be possible up to approximately 2 hours before or after the registered schedule, but vehicle and driver availability cannot be guaranteed.

For detailed driver schedule and availability information, please check the Driver Schedule Dashboard."""

SERVER_MESSAGE_EN = {
    "Tanggal booking tidak boleh tanggal lampau.": "Booking date cannot be in the past.",
    "End date/time must be after start date/time.": "End date/time must be after start date/time.",
    "You have 3 completed trips waiting for review.": "You have 3 completed trips waiting for review.",
    "You already have 3 active or unreviewed bookings.": "You already have 3 active or unreviewed bookings.",
    "Trip not found.": "Trip not found.",
    "You are not allowed to edit this trip.": "You are not allowed to edit this trip.",
    "This booking cannot be canceled.": "This booking cannot be canceled.",
    "Invalid approval action": "Invalid approval action.",
    "You are not authorized to approve this request": "You are not authorized to approve this request.",
    "Request not found": "Request not found.",
    "Requester has no assigned supervisor NIK": "Requester has no assigned supervisor ID.",
    "No request was updated": "No request was updated.",
    "You are not authorized to reject this request.": "You are not authorized to reject this request.",
    "Completed trips cannot be rejected.": "Completed trips cannot be rejected.",
    "Perjalanan belum dapat dimulai.": "Trip cannot be started yet.",
    "Perjalanan sudah diproses atau status berubah.": "Trip has already been processed or its status changed.",
    "Unauthorized": "Unauthorized.",
    "Review not found": "Review not found.",
    "Upload Bukti Tindak lanjut wajib PDF": "Follow-up proof upload must be a PDF.",
    "File wajib dalam bentuk PDF": "File must be a PDF.",
    "Only Super Admin can delete maintenance history.": "Only Super Admin can delete maintenance history.",
    "Kendaraan tidak ditemukan.": "Vehicle not found.",
    "Pilih karyawan terlebih dahulu.": "Please select an employee first.",
    "Karyawan tidak ditemukan.": "Employee not found.",
    "Driver tidak ditemukan.": "Driver not found.",
    "Kendaraan harus dipilih.": "Please select a vehicle.",
    "P2H untuk driver, kendaraan, dan tanggal tersebut sudah ada.": "P2H for this driver, vehicle, and date already exists.",
    "Catatan kerusakan wajib diisi jika ada checklist Tidak OK.": "Damage notes are required when any checklist item is Not OK.",
    "File foto harus JPG, PNG, WEBP, atau PDF.": "Photo file must be JPG, PNG, WEBP, or PDF.",
    "Laporan P2H tidak ditemukan.": "P2H report not found.",
    "File backup tidak ditemukan.": "Backup file not found.",
    "Hanya Super Admin yang boleh restore database.": "Only Super Admin can restore the database.",
    "Hanya Super Admin yang boleh menghapus cadangan database.": "Only Super Admin can delete database backups.",
    "Hari kerja P2H berhasil disimpan": "P2H working days saved successfully.",
    "Only Super Admin can delete all trip history.": "Only Super Admin can delete all trip history.",
    "Please select an Excel file.": "Please select an Excel file.",
    "Invalid employee template columns.": "Invalid employee template columns.",
}

SERVER_MESSAGE_ID = {
    "Booking submitted": "Request booking sudah terkirim.",
    "Trip updated": "Perjalanan berhasil diperbarui.",
    "Booking canceled": "Booking berhasil dibatalkan.",
    "Request approved successfully": "Persetujuan berhasil.",
    "Request rejected successfully": "Request berhasil ditolak.",
    "Request rejected by GA": "Request berhasil ditolak oleh GA.",
    "Trip already started": "Perjalanan sudah dimulai.",
    "Driver update saved": "Update driver berhasil disimpan.",
    "Follow up saved": "Tindak lanjut berhasil disimpan.",
    "Upload Bukti Tindak lanjut berhasil": "Upload bukti tindak lanjut berhasil.",
    "Driver berhasil ditambahkan": "Driver berhasil ditambahkan.",
    "Driver berhasil dihapus": "Driver berhasil dihapus.",
    "Checklist P2H berhasil dikirim": "Checklist P2H berhasil dikirim.",
    "Follow up P2H berhasil disimpan": "Tindak lanjut P2H berhasil disimpan.",
    "Hari kerja P2H berhasil disimpan": "Hari kerja P2H berhasil disimpan.",
    "Hanya Super Admin yang boleh menghapus cadangan database.": "Hanya Super Admin yang boleh menghapus cadangan database.",
}
ASSIGNMENT_BUFFER = timedelta(hours=2)

DEFAULT_POSITION_OPTIONS = [
    "Direktur HRGA",
    "Manager Finance",
    "Manager GA",
    "Manager Marketing",
    "Manager Procurement",
]

DEFAULT_DEPARTMENT_OPTIONS = [
    "Finance",
    "HRGA",
    "Marketing",
    "Operasional",
    "Procurement",
]

MAINTENANCE_REFERENCE_DEFAULTS = {
    "toyota": "https://toyotaastrido.co.id/servis-mobil/service-berkala/",
    "honda": "https://www.honda-indonesia.com/aftersales/costs",
    "suzuki": "https://www.suzuki.co.id/services/automobile",
}


class RowCompat(dict):
    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.values())[key]
        return super().__getitem__(key)


class PgResult:
    def __init__(self, cursor):
        self.cursor = cursor
        self._rows: list[RowCompat] | None = None
        self.lastrowid = None
        self.rowcount = cursor.rowcount

    def _load_rows(self) -> list[RowCompat]:
        if self._rows is None:
            if not self.cursor.description:
                self._rows = []
            else:
                self._rows = [RowCompat(row) for row in self.cursor.fetchall()]
        return self._rows

    def fetchone(self):
        rows_found = self._load_rows()
        return rows_found[0] if rows_found else None

    def fetchall(self):
        return self._load_rows()

    def __iter__(self):
        return iter(self.fetchall())


def pg_sql(query: str) -> str:
    converted = query
    converted = converted.replace("group_concat(er.role, ', ')", "string_agg(er.role, ', ')")
    converted = converted.replace("group_concat(er.role, ',')", "string_agg(er.role, ',')")
    converted = converted.replace("max(coalesce(current_km, 0), ?)", "greatest(coalesce(current_km, 0), ?)")
    converted = converted.replace("insert or ignore into", "insert into")
    converted = converted.replace("insert or replace into app_meta (key, value) values ('seeded', '1')", "insert into app_meta (key, value) values ('seeded', '1') on conflict (key) do update set value = excluded.value")
    converted = converted.replace("insert or replace into app_meta (key, value) values (?, ?)", "insert into app_meta (key, value) values (?, ?) on conflict (key) do update set value = excluded.value")
    if "insert into option_lists" in converted and "on conflict" not in converted.lower():
        converted += " on conflict do nothing"
    if "insert into employee_roles" in converted and "on conflict" not in converted.lower():
        converted += " on conflict do nothing"
    if "insert into employees" in converted and "on conflict" not in converted.lower() and "where not exists" not in converted.lower():
        converted += " on conflict do nothing"
    if "insert into p2h_holidays" in converted and "on conflict" not in converted.lower():
        converted += " on conflict do nothing"
    return converted.replace("?", "%s")


class PgConnection:
    def __init__(self):
        if psycopg2 is None:
            raise RuntimeError("DATABASE_URL is configured, but psycopg2-binary is not installed.")
        self.conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            if exc_type:
                self.conn.rollback()
            else:
                self.conn.commit()
        finally:
            self.conn.close()

    def execute(self, query: str, params: tuple = ()):
        cursor = self.conn.cursor()
        cursor.execute(pg_sql(query), params)
        return PgResult(cursor)

    def executemany(self, query: str, seq_of_params):
        cursor = self.conn.cursor()
        cursor.executemany(pg_sql(query), seq_of_params)
        return PgResult(cursor)

    def executescript(self, script: str) -> None:
        cursor = self.conn.cursor()
        for statement in script.split(";"):
            statement = statement.strip()
            if not statement:
                continue
            statement = statement.replace("integer primary key autoincrement", "serial primary key")
            cursor.execute(pg_sql(statement))

    def commit(self) -> None:
        self.conn.commit()

    def rollback(self) -> None:
        self.conn.rollback()


def db():
    if USE_POSTGRES:
        return PgConnection()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def rows(query: str, params: tuple = ()) -> list[dict]:
    with db() as conn:
        return [dict(row) for row in conn.execute(query, params).fetchall()]


def row(query: str, params: tuple = ()) -> dict | None:
    with db() as conn:
        found = conn.execute(query, params).fetchone()
        return dict(found) if found else None


def execute(query: str, params: tuple = ()) -> None:
    with db() as conn:
        conn.execute(query, params)
        conn.commit()


def wants_json() -> bool:
    return request.headers.get("X-Requested-With") == "fetch" or "application/json" in request.headers.get("Accept", "")


def current_language() -> str:
    return session.get("language", "id")


def ui_message(message: str) -> str:
    if current_language() != "en":
        return SERVER_MESSAGE_ID.get(message, message)
    if message == FULL_BOOKED_ALERT:
        return FULL_BOOKED_ALERT_EN
    for prefix_id, prefix_en in {
        "Backup berhasil dibuat:": "Backup created successfully:",
        "Backup gagal:": "Backup failed:",
        "Restore berhasil dari backup:": "Restore successful from backup:",
        "Restore gagal:": "Restore failed:",
        "File cadangan berhasil dihapus:": "Backup file deleted:",
        "NIK sudah terdaftar:": "Employee ID is already registered:",
        "Invalid role for NIK": "Invalid role for Employee ID",
        "Start KM cannot be lower than previous End KM": "Start KM cannot be lower than previous End KM",
        "End KM cannot be lower than Start KM": "End KM cannot be lower than Start KM",
    }.items():
        if message.startswith(prefix_id):
            return message.replace(prefix_id, prefix_en, 1)
    return SERVER_MESSAGE_EN.get(message, message)


def ok_response(tab: str = "dashboard", message: str = "OK"):
    message = ui_message(message)
    if wants_json():
        return jsonify({"ok": True, "success": True, "message": message})
    return redirect(url_for("index", tab=tab))


def error_response(message: str, tab: str = "dashboard", status_code: int = 400):
    message = ui_message(message)
    if wants_json():
        return jsonify({"ok": False, "success": False, "message": message}), status_code
    return redirect(url_for("index", tab=tab, error=message))


def full_booked_response(status_code: int = 409):
    alert = ui_message(FULL_BOOKED_ALERT)
    if wants_json():
        return jsonify({"ok": False, "success": False, "full_booked": True, "message": alert}), status_code
    return redirect(url_for("index", tab="booking", error=alert, full_booked="1"))


def cutoff_date() -> str:
    return (datetime.now().date() - timedelta(days=92)).isoformat()


def table_columns(conn, table_name: str) -> set[str]:
    if USE_POSTGRES:
        return {
            item["column_name"]
            for item in conn.execute(
                """
                select column_name
                from information_schema.columns
                where table_schema = 'public' and table_name = ?
                """,
                (table_name,),
            ).fetchall()
        }
    return {item["name"] for item in conn.execute(f"pragma table_info({table_name})").fetchall()}


def add_column_if_missing(conn, table_name: str, column_name: str, definition: str) -> None:
    if column_name not in table_columns(conn, table_name):
        conn.execute(f"alter table {table_name} add column {column_name} {definition}")


def list_tables(conn) -> set[str]:
    if USE_POSTGRES:
        return {
            item["table_name"]
            for item in conn.execute(
                """
                select table_name
                from information_schema.tables
                where table_schema = 'public' and table_type = 'BASE TABLE'
                """
            ).fetchall()
        }
    return {item["name"] for item in conn.execute("select name from sqlite_master where type = 'table'").fetchall()}


def remove_p2h_unique_constraint_if_needed(conn) -> None:
    if USE_POSTGRES:
        return
    indexes = conn.execute("pragma index_list(p2h_reports)").fetchall()
    has_unique_constraint = any(item["unique"] and item["origin"] == "u" for item in indexes)
    if not has_unique_constraint:
        return
    if "p2h_reports_unique_backup" in list_tables(conn):
        return
    logger.info("Migrating p2h_reports to allow multiple submissions per driver/vehicle/date")
    conn.executescript(
        """
        create table if not exists p2h_reports_migrated (
            id integer primary key autoincrement,
            report_date text not null,
            submit_time text not null,
            driver_id integer not null,
            vehicle_id integer not null,
            odometer_start integer not null default 0,
            fuel_status text,
            general_note text,
            damage_note text,
            recommendation text,
            status_p2h text not null,
            follow_up_status text,
            follow_up_note text,
            follow_up_action text,
            follow_up_date text,
            created_at text not null,
            updated_at text not null,
            created_by text not null
        );
        insert into p2h_reports_migrated (
            id, report_date, submit_time, driver_id, vehicle_id, odometer_start, fuel_status,
            general_note, damage_note, recommendation, status_p2h, follow_up_status,
            follow_up_note, follow_up_action, follow_up_date, created_at, updated_at, created_by
        )
        select
            id, report_date, submit_time, driver_id, vehicle_id, odometer_start, fuel_status,
            general_note, damage_note, recommendation, status_p2h, follow_up_status,
            follow_up_note, follow_up_action, follow_up_date, created_at, updated_at, created_by
        from p2h_reports;
        alter table p2h_reports rename to p2h_reports_unique_backup;
        alter table p2h_reports_migrated rename to p2h_reports;
        """
    )


def cleanup_old_review_proofs(conn: sqlite3.Connection) -> None:
    cutoff = (datetime.now() - timedelta(days=365)).isoformat(timespec="seconds")
    old_files = conn.execute(
        """
        select id, review_proof_file
        from trip_requests
        where review_proof_uploaded_at is not null
          and review_proof_uploaded_at < ?
        """,
        (cutoff,),
    ).fetchall()
    for item in old_files:
        filename = item["review_proof_file"]
        if filename:
            path = REVIEW_PROOF_DIR / filename
            if path.exists():
                try:
                    path.unlink()
                except OSError:
                    logger.exception("Failed deleting old review proof file id=%s file=%s", item["id"], filename)
        conn.execute(
            """
            update trip_requests
            set review_proof_file = null,
                review_proof_original_name = null,
                review_proof_uploaded_at = null
            where id = ?
            """,
            (item["id"],),
        )


def parse_money(value: str | int | None) -> int:
    if value is None:
        return 0
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return int(digits or 0)


def current_language() -> str:
    return session.get("language", "id")


def status_label(status: str, lang: str | None = None) -> str:
    lang = lang or current_language()
    status = canonical_status(status)
    labels = {
        "id": {
            STATUS_PENDING: "Menunggu Persetujuan Pimpinan",
            STATUS_REJECTED: "Ditolak",
            STATUS_APPROVED: "Disetujui Pimpinan",
            STATUS_PROCESSING: "Diproses GA",
            STATUS_ASSIGNED: "Ditugaskan ke Driver",
            STATUS_ON_TRIP: "Dalam Perjalanan",
            STATUS_COMPLETED: "Selesai",
            STATUS_REVIEWED: "Sudah Direview",
            STATUS_CANCELED: "Dibatalkan",
        },
        "en": {
            STATUS_PENDING: "Pending Leader Approval",
            STATUS_REJECTED: "Rejected",
            STATUS_APPROVED: "Approved by Leader",
            STATUS_PROCESSING: "Processing by GA",
            STATUS_ASSIGNED: "Assigned to Driver",
            STATUS_ON_TRIP: "On Trip",
            STATUS_COMPLETED: "Completed",
            STATUS_REVIEWED: "Reviewed",
            STATUS_CANCELED: "Canceled",
        },
    }
    return labels.get(lang, labels["id"]).get(status, status)


def canonical_status(status: str | None) -> str:
    if not status:
        return ""
    raw = str(status).strip()
    return STATUS_ALIASES.get(raw, STATUS_ALIASES.get(raw.lower(), raw.lower().replace(" ", "_")))


def normalize_roles(raw_roles) -> set[str]:
    if raw_roles is None:
        return set()
    if isinstance(raw_roles, (list, tuple, set)):
        items = raw_roles
    else:
        text = str(raw_roles).strip()
        try:
            parsed = json.loads(text)
            items = parsed if isinstance(parsed, list) else [parsed]
        except Exception:
            items = text.replace(";", ",").split(",")
    return {str(role).strip().lower().replace(" ", "_") for role in items if str(role).strip()}


def option_values(kind: str) -> list[str]:
    return [
        item["value"]
        for item in rows(
            "select value from option_lists where kind = ? order by lower(value)",
            (kind,),
        )
    ]


def add_option_value(kind: str, value: str) -> str:
    cleaned = " ".join(value.strip().split())
    if not cleaned:
        return ""
    execute(
        "insert or ignore into option_lists (kind, value) values (?, ?)",
        (kind, cleaned),
    )
    return cleaned


def redirect_employee_tab():
    return redirect(url_for("index", tab="employees"))


def redirect_employee_error(message: str):
    return redirect(url_for("index", tab="employees", error=message))


def backup_history() -> list[dict]:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    items = []
    backup_files = list(BACKUP_DIR.glob("*.db")) + list(BACKUP_DIR.glob("*.json"))
    for path in sorted(backup_files, key=lambda item: item.stat().st_mtime, reverse=True):
        stat = path.stat()
        items.append({
            "filename": path.name,
            "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            "size_bytes": stat.st_size,
            "size_label": f"{stat.st_size / 1024 / 1024:.2f} MB" if stat.st_size >= 1024 * 1024 else f"{max(1, stat.st_size // 1024)} KB",
        })
    return items


def valid_backup_path(filename: str) -> Path | None:
    safe_name = Path(filename).name
    if safe_name != filename or not safe_name.endswith((".db", ".json")):
        return None
    path = BACKUP_DIR / safe_name
    if not path.exists() or not path.is_file():
        return None
    return path


def create_postgres_json_backup(backup_path: Path) -> None:
    with db() as conn:
        existing_tables = list_tables(conn)
        payload = {
            "database": "postgresql",
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "tables": {},
        }
        for table in APP_TABLES:
            if table in existing_tables:
                payload["tables"][table] = [dict(item) for item in conn.execute(f"select * from {table}").fetchall()]
    backup_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def restore_postgres_json_backup(backup_path: Path) -> None:
    payload = json.loads(backup_path.read_text(encoding="utf-8"))
    table_payload = payload.get("tables") or {}
    with db() as conn:
        existing_tables = list_tables(conn)
        for table in reversed(APP_TABLES):
            if table in existing_tables and table in table_payload:
                conn.execute(f"delete from {table}")
        for table in APP_TABLES:
            rows_to_insert = table_payload.get(table) or []
            if table not in existing_tables or not rows_to_insert:
                continue
            columns = list(rows_to_insert[0].keys())
            placeholders = ", ".join("?" for _ in columns)
            column_sql = ", ".join(columns)
            for item in rows_to_insert:
                conn.execute(
                    f"insert into {table} ({column_sql}) values ({placeholders}) on conflict do nothing",
                    tuple(item.get(column) for column in columns),
                )
        for table in APP_TABLES:
            if table in existing_tables and "id" in table_columns(conn, table):
                conn.execute(
                    "select setval(pg_get_serial_sequence(?, 'id'), coalesce((select max(id) from " + table + "), 1), true)",
                    (table,),
                )


def sync_driver_record(conn: sqlite3.Connection, nik: str, full_name: str, position: str, phone: str, roles: list[str]) -> None:
    existing = conn.execute("select id from drivers where nik = ?", (nik,)).fetchone()
    if "driver" in roles:
        if existing:
            conn.execute(
                "update drivers set driver_name = ?, position = ?, phone = ?, status = 'ACTIVE' where nik = ?",
                (full_name, position, phone, nik),
            )
        else:
            conn.execute(
                "insert into drivers (nik, driver_name, position, phone, status) values (?, ?, ?, ?, 'ACTIVE')",
                (nik, full_name, position, phone),
            )
    elif existing:
        conn.execute("update drivers set status = 'INACTIVE' where nik = ?", (nik,))


def roles_for(nik: str) -> list[str]:
    return [item["role"] for item in rows("select role from employee_roles where nik = ?", (nik,))]


def current_employee() -> dict | None:
    nik = session.get("nik")
    if not nik:
        return None
    emp = row("select * from employees where nik = ?", (nik,))
    if emp:
        emp["roles"] = roles_for(nik)
    return emp


def has_role(emp: dict | None, *roles: str) -> bool:
    if not emp:
        return False
    owned = normalize_roles(emp.get("roles") or emp.get("role"))
    wanted = normalize_roles(roles)
    return "super_admin" in owned or bool(owned.intersection(wanted))


def parse_dt(date_value: str, time_value: str) -> datetime:
    return datetime.strptime(f"{date_value} {time_value}", "%Y-%m-%d %H:%M")


def plate_parity(plate_number: str) -> str | None:
    digits = "".join(ch for ch in plate_number if ch.isdigit())
    if not digits:
        return None
    return "genap" if int(digits[-1]) % 2 == 0 else "ganjil"


def default_maintenance_reference(vehicle_name: str) -> str:
    normalized = (vehicle_name or "").strip().lower()
    for brand, link in MAINTENANCE_REFERENCE_DEFAULTS.items():
        if brand in normalized:
            return link
    return ""


def normalize_plate_rule(value: str | None) -> str:
    value = (value or "bebas").strip().lower()
    return value if value in {"ganjil", "genap", "bebas"} else "bebas"


def overlaps(start_a: datetime, end_a: datetime, start_b: datetime, end_b: datetime) -> bool:
    return start_a < end_b and start_b < end_a


def buffered_range(start_date: str, end_date: str, depart_time: str, return_time: str, apply_buffer: bool = False) -> tuple[datetime, datetime]:
    start = parse_dt(start_date, depart_time)
    end = parse_dt(end_date, return_time)
    if apply_buffer:
        start -= ASSIGNMENT_BUFFER
        end += ASSIGNMENT_BUFFER
    return start, end


def request_with_details(request_id: int) -> dict | None:
    return row(
        """
        select r.*, e.full_name, e.position, e.department, e.supervisor_nik, e.phone,
               s.full_name as supervisor_name, d.driver_name, d.phone as driver_phone,
               v.plate_number, v.vehicle_name, v.vehicle_type, v.current_km as vehicle_current_km
        from trip_requests r
        join employees e on e.nik = r.requester_nik
        left join employees s on s.nik = e.supervisor_nik
        left join drivers d on d.id = r.driver_id
        left join vehicles v on v.id = r.vehicle_id
        where r.id = ?
        """,
        (request_id,),
    )


def trip_rows(where: str = "1=1", params: tuple = (), include_archived: bool = False) -> list[dict]:
    retention_clause = "" if include_archived else " and coalesce(r.end_date, r.travel_date) >= ?"
    final_params = params if include_archived else (*params, cutoff_date())
    return rows(
        f"""
        select r.*, e.full_name, e.position, e.department, e.phone,
               s.full_name as supervisor_name, d.driver_name, d.nik as driver_nik,
               v.plate_number, v.vehicle_name, v.vehicle_type,
               v.current_km as vehicle_current_km
        from trip_requests r
        join employees e on e.nik = r.requester_nik
        left join employees s on s.nik = e.supervisor_nik
        left join drivers d on d.id = r.driver_id
        left join vehicles v on v.id = r.vehicle_id
        where {where}{retention_clause}
        order by r.created_at desc
        """,
        final_params,
    )


def detect_conflict(driver_id: int, vehicle_id: int, start_date: str, end_date: str, depart_time: str, return_time: str, exclude_id: int | None = None) -> str | None:
    new_start = parse_dt(start_date, depart_time)
    new_end = parse_dt(end_date, return_time)
    if new_end <= new_start:
        return "Jam pulang harus lebih besar dari jam berangkat."
    query_start = (new_start - ASSIGNMENT_BUFFER).date().isoformat()
    query_end = (new_end + ASSIGNMENT_BUFFER).date().isoformat()

    existing = rows(
        """
        select r.*, d.driver_name, v.plate_number
        from trip_requests r
        left join drivers d on d.id = r.driver_id
        left join vehicles v on v.id = r.vehicle_id
        where coalesce(r.start_date, r.travel_date) <= ?
          and coalesce(r.end_date, r.travel_date) >= ?
          and r.status in (?, ?, ?)
          and (? is null or r.id != ?)
          and (r.driver_id = ? or r.vehicle_id = ?)
        """,
        (query_end, query_start, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP, exclude_id, exclude_id, driver_id, vehicle_id),
    )
    for item in existing:
        item_start_date = item.get("start_date") or item["travel_date"]
        item_end_date = item.get("end_date") or item["travel_date"]
        item_start, item_end = buffered_range(item_start_date, item_end_date, item["depart_time"], item["return_time"], bool(item.get("driver_id") or item.get("vehicle_id")))
        if overlaps(new_start, new_end, item_start, item_end):
            if item["driver_id"] == driver_id:
                return f"Driver sudah ditugaskan pada jadwal {item['depart_time']} - {item['return_time']} dengan buffer 2 jam sebelum/sesudah."
            return f"Kendaraan {item['plate_number']} sudah dipakai pada jadwal {item['depart_time']} - {item['return_time']} dengan buffer 2 jam sebelum/sesudah."
    return None


def busy_resource_ids(start_date: str, end_date: str, depart_time: str, return_time: str, exclude_id: int | None = None) -> dict:
    new_start = parse_dt(start_date, depart_time)
    new_end = parse_dt(end_date, return_time)
    query_start = (new_start - ASSIGNMENT_BUFFER).date().isoformat()
    query_end = (new_end + ASSIGNMENT_BUFFER).date().isoformat()
    busy_drivers: set[int] = set()
    busy_vehicles: set[int] = set()
    existing = rows(
        """
        select id, driver_id, vehicle_id, start_date, end_date, travel_date, depart_time, return_time
        from trip_requests
        where coalesce(start_date, travel_date) <= ?
          and coalesce(end_date, travel_date) >= ?
          and status in (?, ?, ?)
          and (? is null or id != ?)
        """,
        (query_end, query_start, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP, exclude_id, exclude_id),
    )
    for item in existing:
        item_start, item_end = buffered_range(
            item.get("start_date") or item["travel_date"],
            item.get("end_date") or item["travel_date"],
            item["depart_time"],
            item["return_time"],
            bool(item.get("driver_id") or item.get("vehicle_id")),
        )
        if overlaps(new_start, new_end, item_start, item_end):
            if item.get("driver_id"):
                busy_drivers.add(item["driver_id"])
            if item.get("vehicle_id"):
                busy_vehicles.add(item["vehicle_id"])
    return {"drivers": busy_drivers, "vehicles": busy_vehicles}


def availability_for_trip(item: dict) -> dict:
    busy = busy_resource_ids(
        item.get("start_date") or item["travel_date"],
        item.get("end_date") or item["travel_date"],
        item["depart_time"],
        item["return_time"],
        item["id"],
    )
    drivers = rows("select * from drivers where status = 'ACTIVE' order by driver_name")
    passengers = int(item.get("passengers") or 0)
    plate_rule = normalize_plate_rule(item.get("plate_rule"))
    vehicles = rows("select * from vehicles where status != 'MAINTENANCE' order by plate_number")
    available_vehicles = []
    for vehicle in vehicles:
        if vehicle["id"] in busy["vehicles"] and vehicle["id"] != item.get("vehicle_id"):
            continue
        if passengers and int(vehicle.get("capacity") or 0) < passengers:
            continue
        parity = plate_parity(vehicle.get("plate_number") or "")
        if plate_rule != "bebas" and parity != plate_rule:
            continue
        available_vehicles.append(vehicle)
    available_drivers = [driver for driver in drivers if driver["id"] not in busy["drivers"] or driver["id"] == item.get("driver_id")]
    recommended_vehicle_id = item.get("vehicle_id")
    if not recommended_vehicle_id and available_vehicles:
        recommended_vehicle = sorted(available_vehicles, key=lambda vehicle: (int(vehicle.get("capacity") or 9999), vehicle.get("plate_number") or ""))[0]
        recommended_vehicle_id = recommended_vehicle["id"]
    recommended_driver_id = item.get("driver_id")
    if not recommended_driver_id and recommended_vehicle_id:
        default_driver = next((driver for driver in available_drivers if driver.get("default_vehicle_id") == recommended_vehicle_id), None)
        recommended_driver_id = default_driver["id"] if default_driver else (available_drivers[0]["id"] if available_drivers else None)
    default_vehicle_available = bool(recommended_vehicle_id and any(vehicle["id"] == recommended_vehicle_id for vehicle in available_vehicles))
    return {
        "drivers": available_drivers,
        "vehicles": available_vehicles,
        "default_vehicle_id": recommended_vehicle_id,
        "recommended_vehicle_id": recommended_vehicle_id,
        "recommended_driver_id": recommended_driver_id,
        "default_vehicle_available": default_vehicle_available,
    }


def previous_vehicle_end_km(vehicle_id: int, trip_id: int) -> int:
    item = row(
        """
        select coalesce(km_end, 0) as km_end
        from trip_requests
        where vehicle_id = ?
          and id != ?
          and km_end is not null
        order by coalesce(end_date, travel_date) desc, return_time desc, id desc
        limit 1
        """,
        (vehicle_id, trip_id),
    )
    return int(item["km_end"]) if item else 0


def minimum_vehicle_start_km(vehicle_id: int, trip_id: int) -> int:
    vehicle = row("select coalesce(current_km, 0) as current_km from vehicles where id = ?", (vehicle_id,))
    stored_km = int(vehicle["current_km"]) if vehicle else 0
    return max(stored_km, previous_vehicle_end_km(vehicle_id, trip_id))


def vehicle_matches_request(vehicle_id: int, passengers: int, plate_rule: str) -> str | None:
    vehicle = row("select * from vehicles where id = ?", (vehicle_id,))
    if not vehicle:
        return "Kendaraan tidak ditemukan."
    if vehicle["status"] == "MAINTENANCE":
        return "Kendaraan sedang maintenance."
    if vehicle["status"] not in {"AVAILABLE", "ASSIGNED"}:
        return "Kendaraan tidak tersedia untuk booking umum."
    if int(vehicle.get("capacity") or 0) < int(passengers or 0):
        return f"Kapasitas kursi kendaraan tidak cukup untuk {passengers} penumpang."
    normalized_rule = normalize_plate_rule(plate_rule)
    if normalized_rule != "bebas" and plate_parity(vehicle.get("plate_number") or "") != normalized_rule:
        return f"Nopol kendaraan tidak sesuai pilihan plat {normalized_rule}."
    return None


def available_resources_for_request(start_date: str, end_date: str, depart_time: str, return_time: str, passengers: int, plate_rule: str) -> dict:
    busy = busy_resource_ids(start_date, end_date, depart_time, return_time)
    drivers = rows("select * from drivers where status = 'ACTIVE' order by driver_name")
    vehicles = rows("select * from vehicles where status != 'MAINTENANCE' order by plate_number")
    available_drivers = [driver for driver in drivers if driver["id"] not in busy["drivers"]]
    available_vehicles = []
    normalized_rule = normalize_plate_rule(plate_rule)
    for vehicle in vehicles:
        if vehicle["id"] in busy["vehicles"]:
            continue
        if int(vehicle.get("capacity") or 0) < int(passengers or 0):
            continue
        if normalized_rule != "bebas" and plate_parity(vehicle.get("plate_number") or "") != normalized_rule:
            continue
        available_vehicles.append(vehicle)
    return {"drivers": available_drivers, "vehicles": available_vehicles}


def can_assign_vehicle_pool(requests_to_schedule: list[dict], vehicles: list[dict]) -> bool:
    vehicle_count = len(vehicles)
    matches: dict[int, list[int]] = {}
    for req_index, req in enumerate(requests_to_schedule):
        rule = normalize_plate_rule(req.get("plate_rule"))
        passenger_count = int(req.get("passengers") or 0)
        matches[req_index] = []
        for vehicle_index, vehicle in enumerate(vehicles):
            if int(vehicle.get("capacity") or 0) < passenger_count:
                continue
            if rule != "bebas" and plate_parity(vehicle.get("plate_number") or "") != rule:
                continue
            matches[req_index].append(vehicle_index)
        if not matches[req_index]:
            return False

    assigned_to: dict[int, int] = {}

    def assign(req_index: int, seen: set[int]) -> bool:
        for vehicle_index in matches[req_index]:
            if vehicle_index in seen:
                continue
            seen.add(vehicle_index)
            if vehicle_index not in assigned_to or assign(assigned_to[vehicle_index], seen):
                assigned_to[vehicle_index] = req_index
                return True
        return False

    ordered_requests = sorted(matches, key=lambda item: len(matches[item]))
    if len(ordered_requests) > vehicle_count:
        return False
    return all(assign(req_index, set()) for req_index in ordered_requests)


def booking_capacity_message(start_date: str, end_date: str, depart_time: str, return_time: str, passengers: int, plate_rule: str) -> str | None:
    normalized_rule = normalize_plate_rule(plate_rule)
    drivers = rows("select * from drivers where status = 'ACTIVE'")
    vehicles = rows("select * from vehicles where status != 'MAINTENANCE'")
    matching_vehicles = [
        vehicle for vehicle in vehicles
        if int(vehicle.get("capacity") or 0) >= int(passengers or 0)
        and (normalized_rule == "bebas" or plate_parity(vehicle.get("plate_number") or "") == normalized_rule)
    ]
    if not drivers:
        return "Tidak ada driver aktif yang tersedia. Silakan menghubungi GA jika urgent."
    if not matching_vehicles:
        if normalized_rule != "bebas":
            return f"Tidak ada kendaraan plat {normalized_rule} yang kosong dan sesuai kapasitas {passengers} penumpang. Mohon ubah jadwal/plat atau hubungi GA jika urgent."
        return f"Tidak ada kendaraan kosong yang sesuai kapasitas {passengers} penumpang. Mohon ubah jadwal atau hubungi GA jika urgent."

    new_start = parse_dt(start_date, depart_time)
    new_end = parse_dt(end_date, return_time)
    query_start = (new_start - ASSIGNMENT_BUFFER).date().isoformat()
    query_end = (new_end + ASSIGNMENT_BUFFER).date().isoformat()
    existing = rows(
        """
        select id, start_date, end_date, travel_date, depart_time, return_time,
               passengers, plate_rule, driver_id, vehicle_id, status
        from trip_requests
        where coalesce(start_date, travel_date) <= ?
          and coalesce(end_date, travel_date) >= ?
          and status in (?, ?, ?, ?, ?)
        """,
        (query_end, query_start, STATUS_PENDING, STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP),
    )
    busy_driver_ids: set[int] = set()
    busy_vehicle_ids: set[int] = set()
    queued_requests: list[dict] = []
    for item in existing:
        item_start, item_end = buffered_range(
            item.get("start_date") or item["travel_date"],
            item.get("end_date") or item["travel_date"],
            item["depart_time"],
            item["return_time"],
            bool(item.get("driver_id") or item.get("vehicle_id")),
        )
        if overlaps(new_start, new_end, item_start, item_end):
            if item.get("driver_id"):
                busy_driver_ids.add(item["driver_id"])
            if item.get("vehicle_id"):
                busy_vehicle_ids.add(item["vehicle_id"])
            if not item.get("vehicle_id"):
                queued_requests.append({
                    "passengers": item.get("passengers") or 0,
                    "plate_rule": item.get("plate_rule") or "bebas",
                })
    available_driver_count = len([driver for driver in drivers if driver["id"] not in busy_driver_ids])
    available_vehicles = [vehicle for vehicle in vehicles if vehicle["id"] not in busy_vehicle_ids]
    requests_to_schedule = queued_requests + [{"passengers": passengers, "plate_rule": normalized_rule}]
    if len(requests_to_schedule) > available_driver_count:
        return FULL_BOOKED_ALERT
    if not can_assign_vehicle_pool(requests_to_schedule, available_vehicles):
        return FULL_BOOKED_ALERT
    return None


def seed_data() -> None:
    with db() as conn:
        conn.executescript(
            """
            create table if not exists employees (
                nik text primary key,
                full_name text not null,
                position text not null,
                department text not null,
                supervisor_nik text,
                phone text not null,
                active integer not null default 1
            );
            create table if not exists employee_roles (
                nik text not null,
                role text not null,
                primary key (nik, role)
            );
            create table if not exists drivers (
                id integer primary key autoincrement,
                nik text not null,
                driver_name text not null,
                position text not null,
                phone text not null,
                default_vehicle_id integer,
                sim_expiry_date text,
                status text not null default 'ACTIVE'
            );
            create table if not exists vehicles (
                id integer primary key autoincrement,
                plate_number text not null,
                vehicle_name text not null,
                vehicle_type text not null default 'Operational',
                capacity integer not null,
                status text not null default 'AVAILABLE',
                stnk_expiry_date text,
                kir_expiry_date text,
                current_km integer not null default 0,
                last_maintenance_month text,
                last_service_date text,
                maintenance_km_interval integer not null default 10000,
                maintenance_month_interval integer not null default 6,
                maintenance_reference_url text,
                assigned_to_type text,
                assigned_to_name text,
                assigned_note text
            );
            create table if not exists vehicle_maintenance_history (
                id integer primary key autoincrement,
                vehicle_id integer not null,
                service_date text,
                maintenance_reference_date text,
                km_at_service integer default 0,
                maintenance_types text,
                custom_maintenance_type text,
                parts text,
                custom_part text,
                created_by text,
                created_at text not null
            );
            create table if not exists trip_requests (
                id integer primary key autoincrement,
                request_code text not null,
                requester_nik text not null,
                destination text not null,
                purpose text not null,
                travel_date text not null,
                depart_time text not null,
                return_time text not null,
                passengers integer not null,
                plate_rule text not null default 'bebas',
                notes text,
                status text not null,
                leader_note text,
                ga_note text,
                driver_id integer,
                vehicle_id integer,
                km_start integer,
                km_end integer,
                cost_fuel integer default 0,
                fuel_liters real default 0,
                fuel_type text,
                vehicle_condition_notes text,
                cost_toll integer default 0,
                cost_parking integer default 0,
                rating integer default 0,
                review text,
                review_follow_up text,
                review_proof_file text,
                review_proof_original_name text,
                review_proof_uploaded_at text,
                created_at text not null,
                updated_at text not null
            );
            create table if not exists option_lists (
                kind text not null,
                value text not null,
                primary key (kind, value)
            );
            create table if not exists app_meta (
                key text primary key,
                value text not null
            );
            create table if not exists trip_edit_logs (
                id integer primary key autoincrement,
                trip_id integer not null,
                edited_by text not null,
                action text not null,
                before_data text,
                after_data text,
                created_at text not null
            );
            create table if not exists p2h_reports (
                id integer primary key autoincrement,
                report_date text not null,
                submit_time text not null,
                driver_id integer not null,
                vehicle_id integer not null,
                odometer_start integer not null default 0,
                fuel_status text,
                general_note text,
                damage_note text,
                recommendation text,
                status_p2h text not null,
                follow_up_status text,
                follow_up_note text,
                follow_up_action text,
                follow_up_date text,
                created_at text not null,
                updated_at text not null,
                created_by text not null
            );
            create table if not exists p2h_checklist_items (
                id integer primary key autoincrement,
                p2h_report_id integer not null,
                category text not null,
                item_name text not null,
                result text not null,
                note text
            );
            create table if not exists p2h_attachments (
                id integer primary key autoincrement,
                p2h_report_id integer not null,
                file_path text not null,
                original_name text,
                uploaded_at text not null
            );
            create table if not exists p2h_holidays (
                holiday_date text primary key,
                name text not null
            );
            create table if not exists p2h_workday_overrides (
                id integer primary key autoincrement,
                start_month text not null,
                end_month text not null,
                workdays integer not null,
                updated_by text not null,
                updated_at text not null,
                unique(start_month, end_month)
            );
            """
        )

        add_column_if_missing(conn, "trip_requests", "start_date", "text")
        add_column_if_missing(conn, "trip_requests", "end_date", "text")
        add_column_if_missing(conn, "trip_requests", "map_url", "text")
        add_column_if_missing(conn, "trip_requests", "edited_at", "text")
        add_column_if_missing(conn, "trip_requests", "plate_rule", "text not null default 'bebas'")
        add_column_if_missing(conn, "trip_requests", "fuel_liters", "real default 0")
        add_column_if_missing(conn, "trip_requests", "fuel_type", "text")
        add_column_if_missing(conn, "trip_requests", "vehicle_condition_notes", "text")
        add_column_if_missing(conn, "trip_requests", "cost_parking", "integer default 0")
        add_column_if_missing(conn, "trip_requests", "review_follow_up", "text")
        add_column_if_missing(conn, "trip_requests", "review_proof_file", "text")
        add_column_if_missing(conn, "trip_requests", "review_proof_original_name", "text")
        add_column_if_missing(conn, "trip_requests", "review_proof_uploaded_at", "text")
        add_column_if_missing(conn, "drivers", "default_vehicle_id", "integer")
        add_column_if_missing(conn, "drivers", "sim_expiry_date", "text")
        add_column_if_missing(conn, "vehicles", "vehicle_type", "text not null default 'Operational'")
        add_column_if_missing(conn, "vehicles", "stnk_expiry_date", "text")
        add_column_if_missing(conn, "vehicles", "kir_expiry_date", "text")
        add_column_if_missing(conn, "vehicles", "current_km", "integer not null default 0")
        add_column_if_missing(conn, "vehicles", "last_maintenance_month", "text")
        add_column_if_missing(conn, "vehicles", "last_maintenance_date", "text")
        add_column_if_missing(conn, "vehicles", "last_service_date", "text")
        add_column_if_missing(conn, "vehicles", "maintenance_km_interval", "integer not null default 10000")
        add_column_if_missing(conn, "vehicles", "maintenance_month_interval", "integer not null default 6")
        add_column_if_missing(conn, "vehicles", "maintenance_reference_url", "text")
        add_column_if_missing(conn, "vehicles", "assigned_to_type", "text")
        add_column_if_missing(conn, "vehicles", "assigned_to_name", "text")
        add_column_if_missing(conn, "vehicles", "assigned_note", "text")
        remove_p2h_unique_constraint_if_needed(conn)
        conn.execute("update trip_requests set start_date = travel_date where start_date is null or start_date = ''")
        conn.execute("update trip_requests set end_date = travel_date where end_date is null or end_date = ''")
        REVIEW_PROOF_DIR.mkdir(parents=True, exist_ok=True)
        P2H_ATTACHMENT_DIR.mkdir(parents=True, exist_ok=True)
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        conn.executemany(
            "insert or ignore into p2h_holidays (holiday_date, name) values (?, ?)",
            DEFAULT_P2H_HOLIDAYS.items(),
        )
        cleanup_old_review_proofs(conn)
        for vehicle in conn.execute("select id, vehicle_name from vehicles where maintenance_reference_url is null or maintenance_reference_url = ''").fetchall():
            reference_url = default_maintenance_reference(vehicle["vehicle_name"])
            if reference_url:
                conn.execute("update vehicles set maintenance_reference_url = ? where id = ?", (reference_url, vehicle["id"]))
        conn.execute("update vehicles set last_maintenance_date = last_maintenance_month || '-01' where (last_maintenance_date is null or last_maintenance_date = '') and last_maintenance_month is not null and last_maintenance_month != ''")
        conn.execute("update vehicles set last_service_date = last_maintenance_date where (last_service_date is null or last_service_date = '') and last_maintenance_date is not null and last_maintenance_date != ''")
        status_migrations = {
            STATUS_PENDING: ["PENDING LEADER APPROVAL", "Pending Leader Approval", "Pending", "pending"],
            STATUS_APPROVED: ["APPROVED BY LEADER", "Approved by Leader", "Approved"],
            STATUS_REJECTED: ["REJECTED", "Rejected"],
            STATUS_PROCESSING: ["PROCESSING GA", "Processing GA"],
            STATUS_ASSIGNED: ["ASSIGNED TO DRIVER", "Assigned to Driver", "Assigned"],
            STATUS_ON_TRIP: ["ON TRIP", "On Trip"],
            STATUS_COMPLETED: ["COMPLETED", "Completed"],
            STATUS_REVIEWED: ["REVIEWED", "Reviewed"],
        }
        for canonical, aliases in status_migrations.items():
            conn.execute(
                f"update trip_requests set status = ? where status in ({','.join('?' for _ in aliases)})",
                (canonical, *aliases),
            )
        conn.execute(
            """
            delete from trip_requests
            where status = ?
              and coalesce(rating, 0) > 0
              and coalesce(end_date, travel_date) < ?
            """,
            (STATUS_COMPLETED, cutoff_date()),
        )

        conn.execute("delete from employee_roles where nik in (select nik from employees where active = 0)")
        conn.execute("delete from employees where active = 0")

        for value in DEFAULT_POSITION_OPTIONS:
            conn.execute("insert or ignore into option_lists (kind, value) values ('position', ?)", (value,))
        for value in DEFAULT_DEPARTMENT_OPTIONS:
            conn.execute("insert or ignore into option_lists (kind, value) values ('department', ?)", (value,))

        seeded = conn.execute("select value from app_meta where key = 'seeded'").fetchone()
        employee_count = conn.execute("select count(*) from employees").fetchone()[0]
        if not seeded and employee_count > 0:
            conn.execute("insert or replace into app_meta (key, value) values ('seeded', '1')")
            conn.commit()
            return

        if not seeded:
            employees = [
                ("102145", "BUDI SANTOSO", "STAFF FINANCE", "FINANCE", "102146", "08123456789", ["user"]),
                ("102148", "SITI RAHAYU", "STAFF MARKETING", "MARKETING", "102146", "08125550001", ["user"]),
                ("102146", "IBU RINA", "FINANCE MANAGER", "FINANCE", None, "08129876543", ["pimpinan"]),
                ("102147", "AHMAD FAUZI", "GA ADMIN", "GENERAL AFFAIRS", "900001", "081233344455", ["ga_admin"]),
                ("200145", "BUDI SETIAWAN", "DRIVER OPERASIONAL", "GENERAL AFFAIRS", "102147", "081211122233", ["driver"]),
                ("200146", "ANDI WIJAYA", "DRIVER DIREKSI", "GENERAL AFFAIRS", "102147", "081299988877", ["driver"]),
                ("900001", "DEWI LESTARI", "KEPALA GA", "GENERAL AFFAIRS", None, "081200000001", ["super_admin", "pimpinan", "ga_admin"]),
            ]
            for nik, name, position, department, supervisor, phone, emp_roles in employees:
                conn.execute(
                    """
                    insert or ignore into employees
                    (nik, full_name, position, department, supervisor_nik, phone)
                    values (?, ?, ?, ?, ?, ?)
                    """,
                    (nik, name, position, department, supervisor, phone),
                )
                for role_name in emp_roles:
                    conn.execute("insert or ignore into employee_roles (nik, role) values (?, ?)", (nik, role_name))

            for nik, driver_name, position, phone, status in [
                ("200145", "BUDI SETIAWAN", "DRIVER OPERASIONAL", "081211122233", "ACTIVE"),
                ("200146", "ANDI WIJAYA", "DRIVER DIREKSI", "081299988877", "ACTIVE"),
            ]:
                conn.execute(
                    """
                    insert into drivers (nik, driver_name, position, phone, status)
                    select ?, ?, ?, ?, ?
                    where not exists (select 1 from drivers where nik = ?)
                    """,
                    (nik, driver_name, position, phone, status, nik),
                )

            for plate_number, vehicle_name, vehicle_type, capacity, status in [
                ("B 1234 GA", "Toyota Avanza", "MPV", 6, "AVAILABLE"),
                ("B 5678 GA", "Toyota Innova", "MPV", 7, "AVAILABLE"),
                ("B 9012 GA", "HiAce Commuter", "Van", 14, "AVAILABLE"),
            ]:
                conn.execute(
                    """
                    insert into vehicles (plate_number, vehicle_name, vehicle_type, capacity, status)
                    select ?, ?, ?, ?, ?
                    where not exists (select 1 from vehicles where plate_number = ?)
                    """,
                    (plate_number, vehicle_name, vehicle_type, capacity, status, plate_number),
                )
            conn.execute("insert or replace into app_meta (key, value) values ('seeded', '1')")
        conn.commit()


seed_data()


@app.get("/healthz")
def healthz():
    return jsonify({"ok": True})


@app.get("/")
def index():
    emp = current_employee()
    if not emp:
        return render_template("login.html")
    return render_template(
        "app.html",
        emp=emp,
        role_labels=ROLE_LABELS,
        statuses={
            "pending": STATUS_PENDING,
            "approved": STATUS_APPROVED,
            "processing": STATUS_PROCESSING,
            "assigned": STATUS_ASSIGNED,
            "on_trip": STATUS_ON_TRIP,
            "completed": STATUS_COMPLETED,
            "rejected": STATUS_REJECTED,
        },
    )


@app.post("/login")
def login():
    nik = request.form.get("nik", "").strip()
    session["language"] = request.form.get("language", "id")
    emp = row("select * from employees where nik = ? and active = 1", (nik,))
    if not emp:
        error = "Employee ID was not found or is inactive." if session["language"] == "en" else "NIK tidak ditemukan atau tidak aktif."
        return render_template("login.html", error=error)
    session["nik"] = nik
    return redirect(url_for("index"))


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


@app.post("/requests")
def create_request():
    emp = current_employee()
    if not has_role(emp, "user"):
        return redirect(url_for("index"))

    pending_reviews = row(
        "select count(*) as total from trip_requests where requester_nik = ? and status = ? and coalesce(rating, 0) = 0",
        (emp["nik"], STATUS_COMPLETED),
    )["total"]
    if pending_reviews >= 3:
        return error_response("You have 3 completed trips waiting for review.", "booking")

    active_count = row(
        """
        select count(*) as total
        from trip_requests
        where requester_nik = ?
          and (
            status in (?, ?, ?, ?, ?)
            or (status = ? and coalesce(rating, 0) = 0)
          )
          and coalesce(end_date, travel_date) >= ?
        """,
        (emp["nik"], STATUS_PENDING, STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP, STATUS_COMPLETED, cutoff_date()),
    )["total"]
    if active_count >= 3:
        return error_response("You already have 3 active or unreviewed bookings.", "booking")

    now = datetime.now().isoformat(timespec="seconds")
    start_date = request.form.get("start_date") or request.form.get("travel_date")
    end_date = request.form.get("end_date") or start_date
    if start_date < datetime.now().date().isoformat():
        return error_response("Tanggal booking tidak boleh tanggal lampau.", "booking")
    if parse_dt(end_date, request.form["return_time"]) <= parse_dt(start_date, request.form["depart_time"]):
        return error_response("End date/time must be after start date/time.", "booking")
    passengers = int(request.form["passengers"])
    plate_rule = normalize_plate_rule(request.form.get("plate_rule"))
    capacity_message = booking_capacity_message(
        start_date,
        end_date,
        request.form["depart_time"],
        request.form["return_time"],
        passengers,
        plate_rule,
    )
    if capacity_message:
        if capacity_message == FULL_BOOKED_ALERT:
            return full_booked_response()
        return error_response(capacity_message, "booking")
    availability = available_resources_for_request(
        start_date,
        end_date,
        request.form["depart_time"],
        request.form["return_time"],
        passengers,
        plate_rule,
    )
    if not availability["drivers"] or not availability["vehicles"]:
        return full_booked_response()
    with db() as conn:
        conn.execute(
            """
            insert into trip_requests (
                request_code, requester_nik, destination, purpose, travel_date, start_date, end_date,
                depart_time, return_time, passengers, plate_rule, notes, map_url, status, created_at, updated_at
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "REQ-" + datetime.now().strftime("%Y%m%d%H%M%S"),
                emp["nik"],
                request.form["destination"].strip(),
                request.form["purpose"].strip(),
                start_date,
                start_date,
                end_date,
                request.form["depart_time"],
                request.form["return_time"],
                passengers,
                plate_rule,
                request.form.get("notes", "").strip(),
                request.form.get("map_url", "").strip(),
                STATUS_PENDING,
                now,
                now,
            ),
        )
        conn.commit()
    if wants_json():
        return ok_response("booking", "Booking submitted")
    return redirect(url_for("index", tab="booking", success="Request booking sudah terkirim."))


@app.post("/requests/<int:request_id>/update")
def update_request(request_id: int):
    emp = current_employee()
    item = request_with_details(request_id)
    if not emp or not item:
        return error_response("Trip not found.", "history", 404)
    can_user_edit = item["requester_nik"] == emp["nik"] and item["status"] in EDITABLE_BOOKING_STATUSES
    can_ga_edit = has_role(emp, "ga_admin")
    if not can_user_edit and not can_ga_edit:
        return error_response("You are not allowed to edit this trip.", "history", 403)
    start_date = request.form["start_date"]
    end_date = request.form["end_date"]
    if can_user_edit and start_date < datetime.now().date().isoformat():
        return error_response("Tanggal booking tidak boleh tanggal lampau.", "history")
    if parse_dt(end_date, request.form["return_time"]) <= parse_dt(start_date, request.form["depart_time"]):
        return error_response("End date/time must be after start date/time.", "history")
    driver_id = request.form.get("driver_id") or item.get("driver_id")
    vehicle_id = request.form.get("vehicle_id") or item.get("vehicle_id")
    if can_ga_edit and driver_id and vehicle_id:
        vehicle_error = vehicle_matches_request(
            int(vehicle_id),
            int(request.form["passengers"]),
            normalize_plate_rule(request.form.get("plate_rule") or item.get("plate_rule")),
        )
        if vehicle_error:
            return error_response(vehicle_error, "ga")
        conflict = detect_conflict(int(driver_id), int(vehicle_id), start_date, end_date, request.form["depart_time"], request.form["return_time"], request_id)
        if conflict:
            return error_response(conflict, "ga")
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute(
            """
            update trip_requests
            set destination = ?, purpose = ?, travel_date = ?, start_date = ?, end_date = ?,
                depart_time = ?, return_time = ?, passengers = ?, plate_rule = ?, notes = ?, map_url = ?,
                driver_id = coalesce(?, driver_id), vehicle_id = coalesce(?, vehicle_id),
                edited_at = ?, updated_at = ?
            where id = ?
            """,
            (
                request.form["destination"].strip(),
                request.form["purpose"].strip(),
                start_date,
                start_date,
                end_date,
                request.form["depart_time"],
                request.form["return_time"],
                int(request.form["passengers"]),
                normalize_plate_rule(request.form.get("plate_rule") or item.get("plate_rule")),
                request.form.get("notes", "").strip(),
                request.form.get("map_url", "").strip(),
                int(driver_id) if can_ga_edit and driver_id else None,
                int(vehicle_id) if can_ga_edit and vehicle_id else None,
                now,
                now,
                request_id,
            ),
        )
        conn.execute(
            "insert into trip_edit_logs (trip_id, edited_by, action, before_data, after_data, created_at) values (?, ?, ?, ?, ?, ?)",
            (request_id, emp["nik"], "GA_EDIT" if can_ga_edit else "USER_EDIT", str(dict(item)), str(dict(request.form)), now),
        )
        conn.commit()
    return ok_response("ga" if can_ga_edit else "history", "Trip updated")


@app.post("/requests/<int:request_id>/cancel")
def cancel_request(request_id: int):
    emp = current_employee()
    item = request_with_details(request_id)
    if not emp or not item or item["requester_nik"] != emp["nik"]:
        return error_response("Trip not found.", "history", 404)
    if item["status"] in {STATUS_ASSIGNED, STATUS_ON_TRIP, STATUS_COMPLETED, STATUS_REVIEWED, STATUS_CANCELED}:
        return error_response("This booking cannot be canceled.", "history", 409)
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute(
            "update trip_requests set status = ?, updated_at = ? where id = ?",
            (STATUS_CANCELED, now, request_id),
        )
        conn.execute(
            "insert into trip_edit_logs (trip_id, edited_by, action, before_data, after_data, created_at) values (?, ?, ?, ?, ?, ?)",
            (request_id, emp["nik"], "USER_CANCEL", str(dict(item)), STATUS_CANCELED, now),
        )
        conn.commit()
    return ok_response("history", "Booking canceled")


@app.post("/requests/<int:request_id>/approval")
def approval(request_id: int):
    emp = current_employee()

    data = request.get_json(silent=True) or request.form
    action = (data.get("action") or "").strip().lower()
    leader_note = (data.get("leader_note") or "").strip()

    logger.debug(
        "Approval request received booking_id=%s leader=%s roles=%s action=%s",
        request_id,
        emp["nik"] if emp else None,
        emp.get("roles") if emp else None,
        action,
    )

    if action not in {"approve", "reject"}:
        logger.debug(
            "Approval failed invalid action booking_id=%s action=%s",
            request_id,
            action,
        )
        return error_response("Invalid approval action", "approval", 400)

    if not emp or not has_role(emp, "pimpinan"):
        logger.debug(
            "Approval failed unauthorized booking_id=%s user=%s",
            request_id,
            emp["nik"] if emp else None,
        )
        return error_response("You are not authorized to approve this request", "approval", 403)

    try:
        with db() as conn:
            item_row = conn.execute(
                """
                SELECT r.*, e.supervisor_nik, e.full_name as requester_name, e.supervisor_nik as requester_supervisor_nik
                FROM trip_requests r
                JOIN employees e ON e.nik = r.requester_nik
                WHERE r.id = ?
                """,
                (request_id,),
            ).fetchone()

            if not item_row:
                logger.debug(
                    "Approval failed missing booking_id=%s leader=%s",
                    request_id,
                    emp["nik"],
                )
                return error_response("Request not found", "approval", 404)

            item = dict(item_row)
            old_status = item.get("status")
            canonical_old_status = canonical_status(old_status)

            allowed_pending_statuses = {
                STATUS_PENDING,
                "Pending",
                "Pending Leader Approval",
                "PENDING LEADER APPROVAL",
            }

            logger.debug(
                "Approval context booking_id=%s action=%s leader=%s roles=%s requester=%s supervisor_nik=%s old_status=%s canonical_status=%s",
                request_id,
                action,
                emp["nik"],
                emp.get("roles"),
                item.get("requester_nik"),
                item.get("supervisor_nik"),
                old_status,
                canonical_old_status,
            )

            if canonical_old_status not in {canonical_status(status) for status in allowed_pending_statuses}:
                logger.debug(
                    "Approval failed invalid status booking_id=%s status=%s leader=%s",
                    request_id,
                    old_status,
                    emp["nik"],
                )
                return error_response(
                    f"Request status is not pending approval. Current status: {old_status}",
                    "approval",
                    409,
                )

            supervisor_nik = (item.get("supervisor_nik") or "").strip()
            if not supervisor_nik and not has_role(emp, "super_admin"):
                logger.debug(
                    "Approval failed missing supervisor_nik booking_id=%s leader=%s requester=%s",
                    request_id,
                    emp["nik"],
                    item.get("requester_nik"),
                )
                return error_response("Requester has no assigned supervisor NIK", "approval", 400)

            if emp["nik"] != supervisor_nik and not has_role(emp, "super_admin"):
                logger.debug(
                    "Approval failed wrong leader booking_id=%s leader=%s supervisor=%s",
                    request_id,
                    emp["nik"],
                    supervisor_nik,
                )
                return error_response(
                    "You are not the assigned leader for this request",
                    "approval",
                    403,
                )

            status = STATUS_APPROVED if action == "approve" else STATUS_REJECTED

            result = conn.execute(
                """
                UPDATE trip_requests
                SET status = ?, leader_note = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    status,
                    leader_note,
                    datetime.now().isoformat(timespec="seconds"),
                    request_id,
                ),
            )

            if result.rowcount != 1:
                conn.rollback()
                logger.debug(
                    "Approval update affected no rows booking_id=%s leader=%s action=%s",
                    request_id,
                    emp["nik"],
                    action,
                )
                return error_response("No request was updated", "approval", 409)

            conn.commit()

            logger.debug(
                "Approval update committed booking_id=%s leader=%s requester=%s supervisor_nik=%s action=%s old_status=%s new_status=%s rowcount=%s",
                request_id,
                emp["nik"],
                item.get("requester_nik"),
                supervisor_nik,
                action,
                old_status,
                status,
                result.rowcount,
            )

    except Exception as e:
        logger.exception(
            "Approval error booking_id=%s leader=%s action=%s",
            request_id,
            emp["nik"] if emp else None,
            action,
        )
        return error_response(f"Approval error: {str(e)}", "approval", 500)

    message = "Request approved successfully" if action == "approve" else "Request rejected successfully"
    return ok_response("approval", message)


@app.post("/requests/<int:request_id>/assign")
def assign(request_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    item = request_with_details(request_id)
    if not item:
        return redirect(url_for("index"))
    driver_id = int(request.form["driver_id"])
    vehicle_id = int(request.form["vehicle_id"])
    vehicle_error = vehicle_matches_request(vehicle_id, int(item["passengers"]), item.get("plate_rule"))
    if vehicle_error:
        return error_response(vehicle_error, "ga")
    conflict = detect_conflict(driver_id, vehicle_id, item.get("start_date") or item["travel_date"], item.get("end_date") or item["travel_date"], item["depart_time"], item["return_time"], request_id)
    if conflict:
        return error_response(conflict, "ga")
    execute(
        "update trip_requests set driver_id = ?, vehicle_id = ?, ga_note = ?, status = ?, updated_at = ? where id = ?",
        (driver_id, vehicle_id, request.form.get("ga_note", "").strip(), STATUS_ASSIGNED, datetime.now().isoformat(timespec="seconds"), request_id),
    )
    return ok_response("ga", "Assignment saved")


@app.post("/requests/<int:request_id>/ga-reject")
def ga_reject(request_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return error_response("You are not authorized to reject this request.", "ga", 403)
    item = request_with_details(request_id)
    if not item:
        return error_response("Trip not found.", "ga", 404)
    if item["status"] in {STATUS_COMPLETED, STATUS_REVIEWED}:
        return error_response("Completed trips cannot be rejected.", "ga", 409)
    now = datetime.now().isoformat(timespec="seconds")
    note = request.form.get("ga_note", "").strip()
    with db() as conn:
        conn.execute(
            "update trip_requests set status = ?, ga_note = ?, updated_at = ? where id = ?",
            (STATUS_REJECTED, note, now, request_id),
        )
        conn.execute(
            "insert into trip_edit_logs (trip_id, edited_by, action, before_data, after_data, created_at) values (?, ?, ?, ?, ?, ?)",
            (request_id, emp["nik"], "GA_REJECT", str(dict(item)), note, now),
        )
        conn.commit()
    return ok_response("ga", "Request rejected by GA")


@app.post("/requests/<int:request_id>/driver")
def driver_update(request_id: int):
    emp = current_employee()
    item = request_with_details(request_id)
    if not item or not has_role(emp, "driver"):
        return redirect(url_for("index"))
    driver = row("select * from drivers where id = ? and nik = ?", (item["driver_id"], emp["nik"]))
    if not driver and not has_role(emp, "super_admin"):
        return redirect(url_for("index"))
    action = request.form.get("action")
    if action == "start":
        minimum_km = minimum_vehicle_start_km(item["vehicle_id"], request_id) if item.get("vehicle_id") else 0
        km_start = int(request.form.get("km_start") or minimum_km or 0)
        if km_start < minimum_km:
            return error_response(f"Start KM cannot be lower than previous End KM ({minimum_km}).", "driver")
        if item["status"] == STATUS_ON_TRIP:
            return ok_response("driver", "Trip already started")
        if item["status"] not in {STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED}:
            return error_response("Perjalanan belum dapat dimulai.", "driver", 409)
        with db() as conn:
            result = conn.execute(
                """
                update trip_requests
                set status = ?, km_start = ?, updated_at = ?
                where id = ? and status in (?, ?, ?)
                """,
                (STATUS_ON_TRIP, km_start, datetime.now().isoformat(timespec="seconds"), request_id, STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED),
            )
            conn.commit()
        if result.rowcount == 0:
            return error_response("Perjalanan sudah diproses atau status berubah.", "driver", 409)
    elif action == "finish":
        km_end = int(request.form.get("km_end") or 0)
        km_start = int(item.get("km_start") or 0)
        if km_end < km_start:
            return error_response(f"End KM cannot be lower than Start KM ({km_start}).", "driver")
        with db() as conn:
            now = datetime.now().isoformat(timespec="seconds")
            conn.execute(
                """
                update trip_requests
                set status = ?, km_end = ?, cost_fuel = ?, fuel_liters = ?, fuel_type = ?,
                    vehicle_condition_notes = ?, cost_toll = ?, cost_parking = ?, updated_at = ?
                where id = ?
                """,
                (
                    STATUS_COMPLETED,
                    km_end,
                    parse_money(request.form.get("cost_fuel")),
                    float(request.form.get("fuel_liters") or 0),
                    request.form.get("fuel_type", "").strip(),
                    request.form.get("vehicle_condition_notes", "").strip(),
                    parse_money(request.form.get("cost_toll")),
                    parse_money(request.form.get("cost_parking")),
                    now,
                    request_id,
                ),
            )
            if item.get("vehicle_id"):
                conn.execute(
                    "update vehicles set current_km = max(coalesce(current_km, 0), ?) where id = ?",
                    (km_end, item["vehicle_id"]),
                )
            conn.commit()
    return ok_response("driver", "Driver update saved")


@app.post("/requests/<int:request_id>/review")
def review(request_id: int):
    emp = current_employee()
    item = request_with_details(request_id)
    if not emp or not item or item["requester_nik"] != emp["nik"]:
        return redirect(url_for("index"))
    rating = int(request.form["rating"])
    execute(
        "update trip_requests set rating = ?, review = ?, updated_at = ? where id = ?",
        (rating, request.form.get("review", "").strip(), datetime.now().isoformat(timespec="seconds"), request_id),
    )
    return redirect(url_for("index", tab="history"))


@app.post("/requests/<int:request_id>/review-follow-up")
def update_review_follow_up(request_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return error_response("Unauthorized", "user-reviews", 403)
    item = request_with_details(request_id)
    if not item or not item.get("rating"):
        return error_response("Review not found", "user-reviews", 404)
    execute(
        "update trip_requests set review_follow_up = ? where id = ?",
        (request.form.get("review_follow_up", "").strip(), request_id),
    )
    return ok_response("user-reviews", "Follow up saved")


@app.post("/requests/<int:request_id>/review-proof")
def upload_review_proof(request_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return error_response("Unauthorized", "user-reviews", 403)
    item = request_with_details(request_id)
    if not item or not item.get("rating"):
        return error_response("Review not found", "user-reviews", 404)
    uploaded = request.files.get("review_proof")
    if not uploaded or not uploaded.filename:
        return error_response("Upload Bukti Tindak lanjut wajib PDF", "user-reviews")
    original_name = secure_filename(uploaded.filename)
    if not original_name.lower().endswith(".pdf") or uploaded.mimetype not in {"application/pdf", "application/x-pdf", ""}:
        return error_response("File wajib dalam bentuk PDF", "user-reviews")
    REVIEW_PROOF_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"review_{request_id}_{timestamp}.pdf"
    existing_file = item.get("review_proof_file")
    if existing_file:
        old_path = REVIEW_PROOF_DIR / existing_file
        if old_path.exists():
            old_path.unlink()
    uploaded.save(REVIEW_PROOF_DIR / filename)
    execute(
        """
        update trip_requests
        set review_proof_file = ?,
            review_proof_original_name = ?,
            review_proof_uploaded_at = ?
        where id = ?
        """,
        (filename, original_name, datetime.now().isoformat(timespec="seconds"), request_id),
    )
    return ok_response("user-reviews", "Upload Bukti Tindak lanjut berhasil")


@app.get("/requests/<int:request_id>/review-proof")
def download_review_proof(request_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    item = request_with_details(request_id)
    filename = item.get("review_proof_file") if item else None
    if not filename or not (REVIEW_PROOF_DIR / filename).exists():
        return redirect(url_for("index", tab="user-reviews", error="Bukti tindak lanjut belum tersedia"))
    return send_from_directory(
        REVIEW_PROOF_DIR,
        filename,
        as_attachment=False,
        download_name=item.get("review_proof_original_name") or filename,
        mimetype="application/pdf",
    )


@app.post("/employees")
def add_employee():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    nik = request.form["nik"].strip()
    if row("select nik from employees where nik = ?", (nik,)):
        logger.debug("Employee insert blocked; duplicate NIK=%s", nik)
        return redirect_employee_error("NIK sudah terdaftar")

    roles = request.form.getlist("roles")
    position = add_option_value("position", request.form["position"])
    department = add_option_value("department", request.form["department"])
    full_name = request.form["full_name"].strip().upper()
    phone = request.form["phone"].strip()
    with db() as conn:
        try:
            logger.debug("Inserting employee NIK=%s name=%s roles=%s", nik, full_name, roles or ["user"])
            conn.execute(
                "insert into employees (nik, full_name, position, department, supervisor_nik, phone) values (?, ?, ?, ?, ?, ?)",
                (
                    nik,
                    full_name,
                    position,
                    department,
                    request.form.get("supervisor_nik") or None,
                    phone,
                ),
            )
            for role_name in roles or ["user"]:
                conn.execute("insert into employee_roles (nik, role) values (?, ?)", (nik, role_name))
            sync_driver_record(conn, nik, full_name, position, phone, roles)
            conn.commit()
        except sqlite3.IntegrityError:
            conn.rollback()
            logger.exception("Employee insert failed due to duplicate or constraint issue. NIK=%s", nik)
            return redirect_employee_error("NIK sudah terdaftar")
    return redirect_employee_tab()


@app.post("/employees/<nik>/update")
def update_employee(nik: str):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    roles = request.form.getlist("roles") or ["user"]
    position = add_option_value("position", request.form["position"])
    department = add_option_value("department", request.form["department"])
    full_name = request.form["full_name"].strip().upper()
    phone = request.form["phone"].strip()
    with db() as conn:
        conn.execute(
            """
            update employees
            set full_name = ?, position = ?, department = ?, supervisor_nik = ?, phone = ?, active = 1
            where nik = ?
            """,
            (full_name, position, department, request.form.get("supervisor_nik") or None, phone, nik),
        )
        conn.execute("delete from employee_roles where nik = ?", (nik,))
        for role_name in roles:
            conn.execute("insert into employee_roles (nik, role) values (?, ?)", (nik, role_name))
        sync_driver_record(conn, nik, full_name, position, phone, roles)
        conn.commit()
    return redirect_employee_tab()


@app.post("/employees/<nik>/delete")
def delete_employee(nik: str):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    if emp["nik"] == nik:
        return redirect_employee_error("Tidak bisa menghapus user yang sedang login.")
    with db() as conn:
        logger.debug("Deleting employee permanently NIK=%s", nik)
        conn.execute("delete from employee_roles where nik = ?", (nik,))
        conn.execute("delete from employees where nik = ?", (nik,))
        conn.execute("update drivers set status = 'INACTIVE' where nik = ?", (nik,))
        conn.commit()
        remaining = conn.execute("select count(*) from employees where nik = ?", (nik,)).fetchone()[0]
        logger.debug("Employee delete complete NIK=%s remaining_rows=%s", nik, remaining)
    return redirect_employee_tab()


@app.post("/options")
def add_option():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return jsonify({"ok": False, "message": "Tidak berwenang."}), 403
    payload = request.get_json(silent=True) or {}
    kind = payload.get("kind", "")
    if kind not in {"position", "department"}:
        return jsonify({"ok": False, "message": "Jenis dropdown tidak valid."}), 400
    value = add_option_value(kind, payload.get("value", ""))
    if not value:
        return jsonify({"ok": False, "message": "Nilai tidak boleh kosong."}), 400
    return jsonify({"ok": True, "value": value, "options": option_values(kind)})


@app.post("/options/delete")
def delete_option():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return jsonify({"ok": False, "message": "Tidak berwenang."}), 403
    payload = request.get_json(silent=True) or {}
    kind = payload.get("kind", "")
    value = payload.get("value", "").strip()
    if kind not in {"position", "department"} or not value:
        return jsonify({"ok": False, "message": "Pilihan tidak valid."}), 400
    execute("delete from option_lists where kind = ? and value = ?", (kind, value))
    return jsonify({"ok": True, "options": option_values(kind)})


@app.post("/vehicles")
def add_vehicle():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    status = request.form["status"]
    execute(
        """
        insert into vehicles (
            vehicle_name, plate_number, vehicle_type, capacity, status, stnk_expiry_date, kir_expiry_date,
            current_km, last_service_date, last_maintenance_month, last_maintenance_date,
            maintenance_km_interval, maintenance_month_interval, maintenance_reference_url
        )
        values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            request.form["vehicle_name"].strip(),
            request.form["plate_number"].strip().upper(),
            request.form["vehicle_type"].strip(),
            int(request.form["capacity"]),
            status,
            request.form.get("stnk_expiry_date", "").strip() or None,
            request.form.get("kir_expiry_date", "").strip() or None,
            int(request.form.get("current_km") or 0),
            request.form.get("last_service_date", "").strip() or None,
            (request.form.get("last_maintenance_date", "").strip() or request.form.get("last_service_date", "").strip() or "")[:7] or None,
            request.form.get("last_maintenance_date", "").strip() or request.form.get("last_service_date", "").strip() or None,
            int(request.form.get("maintenance_km_interval") or 10000),
            int(request.form.get("maintenance_month_interval") or 6),
            request.form.get("maintenance_reference_url", "").strip() or default_maintenance_reference(request.form["vehicle_name"]),
        ),
    )
    return redirect(url_for("index", tab="vehicles"))


@app.post("/vehicles/<int:vehicle_id>/update")
def update_vehicle(vehicle_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    status = request.form["status"]
    execute(
        """
        update vehicles
        set vehicle_name = ?, plate_number = ?, vehicle_type = ?, capacity = ?, status = ?,
            stnk_expiry_date = ?, kir_expiry_date = ?
        where id = ?
        """,
        (
            request.form["vehicle_name"].strip(),
            request.form["plate_number"].strip().upper(),
            request.form["vehicle_type"].strip(),
            int(request.form["capacity"]),
            status,
            request.form.get("stnk_expiry_date", "").strip() or None,
            request.form.get("kir_expiry_date", "").strip() or None,
            vehicle_id,
        ),
    )
    return redirect(url_for("index", tab="vehicles"))


@app.post("/vehicles/<int:vehicle_id>/maintenance")
def update_vehicle_maintenance(vehicle_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    current_km_value = request.form.get("current_km")
    can_edit_km = has_role(emp, "super_admin") and current_km_value not in {None, ""}
    is_super_admin = has_role(emp, "super_admin")
    existing = row("select * from vehicles where id = ?", (vehicle_id,))
    if not existing:
        return redirect(url_for("index", tab="maintenance"))
    service_date = request.form.get("last_service_date", "").strip() or None
    maintenance_date = request.form.get("last_maintenance_date", "").strip() or service_date
    maintenance_month = maintenance_date[:7] if maintenance_date else None
    month_interval = int(request.form.get("maintenance_month_interval") or existing.get("maintenance_month_interval") or 6) if is_super_admin else int(existing.get("maintenance_month_interval") or 6)
    km_interval = int(request.form.get("maintenance_km_interval") or existing.get("maintenance_km_interval") or 10000)
    maintenance_types = [value for value in request.form.getlist("maintenance_types") if value in MAINTENANCE_TYPE_OPTIONS]
    maintenance_parts = [value for value in request.form.getlist("maintenance_parts") if value in MAINTENANCE_PART_OPTIONS]
    custom_maintenance_type = request.form.get("custom_maintenance_type", "").strip()
    custom_part = request.form.get("custom_part", "").strip()
    has_realization = bool(maintenance_types or maintenance_parts or custom_maintenance_type or custom_part)
    km_at_service = int(current_km_value or existing.get("current_km") or 0)
    with db() as conn:
        if can_edit_km:
            conn.execute(
                """
                update vehicles
                set current_km = ?, last_service_date = ?, last_maintenance_date = ?, last_maintenance_month = ?, maintenance_km_interval = ?,
                    maintenance_month_interval = ?, maintenance_reference_url = ?
                where id = ?
                """,
                (
                    int(current_km_value),
                    service_date,
                    maintenance_date,
                    maintenance_month,
                    km_interval,
                    month_interval,
                    request.form.get("maintenance_reference_url", "").strip(),
                    vehicle_id,
                ),
            )
        else:
            conn.execute(
                """
                update vehicles
                set last_service_date = ?, last_maintenance_date = ?, last_maintenance_month = ?, maintenance_km_interval = ?,
                    maintenance_month_interval = ?, maintenance_reference_url = ?
                where id = ?
                """,
                (
                    service_date,
                    maintenance_date,
                    maintenance_month,
                    km_interval,
                    month_interval,
                    request.form.get("maintenance_reference_url", "").strip(),
                    vehicle_id,
                ),
            )
        if has_realization:
            conn.execute(
                """
                insert into vehicle_maintenance_history (
                    vehicle_id, service_date, maintenance_reference_date, km_at_service,
                    maintenance_types, custom_maintenance_type, parts, custom_part, created_by, created_at
                )
                values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    vehicle_id,
                    service_date,
                    maintenance_date,
                    km_at_service,
                    json.dumps(maintenance_types, ensure_ascii=False),
                    custom_maintenance_type,
                    json.dumps(maintenance_parts, ensure_ascii=False),
                    custom_part,
                    emp["nik"],
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
        conn.commit()
    return redirect(url_for("index", tab="maintenance"))


@app.post("/maintenance-history/<int:history_id>/delete")
def delete_maintenance_history(history_id: int):
    emp = current_employee()
    if not has_role(emp, "super_admin"):
        return error_response("Only Super Admin can delete maintenance history.", "maintenance", 403)
    execute("delete from vehicle_maintenance_history where id = ?", (history_id,))
    return ok_response("maintenance", "Maintenance history deleted")


@app.post("/drivers/<int:driver_id>/default-vehicle")
def update_driver_default_vehicle(driver_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    default_vehicle_id = request.form.get("default_vehicle_id") or None
    sim_expiry_date = request.form.get("sim_expiry_date", "").strip() or None
    if default_vehicle_id and not row("select id from vehicles where id = ?", (default_vehicle_id,)):
        return error_response("Kendaraan tidak ditemukan.", "vehicles", 404)
    execute(
        "update drivers set default_vehicle_id = ?, sim_expiry_date = ? where id = ?",
        (int(default_vehicle_id) if default_vehicle_id else None, sim_expiry_date, driver_id),
    )
    return ok_response("vehicles", "Default vehicle saved")


@app.post("/drivers/add-registered")
def add_registered_driver():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    nik = request.form.get("nik", "").strip()
    if not nik:
        return error_response("Pilih karyawan terlebih dahulu.", "vehicles", 400)
    employee = row("select * from employees where nik = ? and active = 1", (nik,))
    if not employee:
        return error_response("Karyawan tidak ditemukan.", "vehicles", 404)
    roles = roles_for(nik)
    if "driver" not in roles:
        roles.append("driver")
    with db() as conn:
        conn.execute("insert or ignore into employee_roles (nik, role) values (?, 'driver')", (nik,))
        sync_driver_record(
            conn,
            nik,
            employee["full_name"],
            employee["position"],
            employee["phone"],
            roles,
        )
        conn.commit()
    app.logger.info("Driver added from registered employee nik=%s by=%s", nik, emp.get("nik"))
    return ok_response("vehicles", "Driver berhasil ditambahkan")


@app.post("/drivers/<int:driver_id>/delete")
def delete_driver(driver_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    driver = row("select * from drivers where id = ?", (driver_id,))
    if not driver:
        return error_response("Driver tidak ditemukan.", "vehicles", 404)
    with db() as conn:
        conn.execute("update drivers set status = 'INACTIVE', default_vehicle_id = null where id = ?", (driver_id,))
        conn.execute("delete from employee_roles where nik = ? and role = 'driver'", (driver["nik"],))
        conn.commit()
    app.logger.info("Driver deactivated id=%s nik=%s by=%s", driver_id, driver["nik"], emp.get("nik"))
    return ok_response("vehicles", "Driver berhasil dihapus")


@app.post("/p2h")
def submit_p2h():
    emp = current_employee()
    if not has_role(emp, "driver"):
        return redirect(url_for("index"))
    driver = row("select * from drivers where nik = ? and status = 'ACTIVE'", (emp["nik"],))
    if not driver:
        return error_response("Driver tidak ditemukan.", "p2h-checklist", 404)
    report_date = request.form.get("report_date", today_iso()).strip() or today_iso()
    vehicle_id = request.form.get("vehicle_id", "").strip()
    if not vehicle_id or not row("select id from vehicles where id = ?", (vehicle_id,)):
        return error_response("Kendaraan harus dipilih.", "p2h-checklist", 400)
    odometer_start = int(request.form.get("odometer_start") or 0)
    fuel_status = request.form.get("fuel_status", "").strip()
    general_note = request.form.get("general_note", "").strip()
    damage_note = request.form.get("damage_note", "").strip()
    recommendation = request.form.get("recommendation", "").strip()
    checklist_payload = []
    has_not_ok = False
    for category, items in P2H_CHECKLIST.items():
        for item in items:
            key = f"p2h__{category}__{item}"
            result = request.form.get(key, "OK").strip()
            if result not in P2H_RESULTS:
                result = "OK"
            note = request.form.get(key + "__note", "").strip()
            if result == "Tidak OK":
                has_not_ok = True
            checklist_payload.append((category, item, result, note))
    if has_not_ok and not damage_note:
        return error_response("Catatan kerusakan wajib diisi jika ada checklist Tidak OK.", "p2h-checklist", 400)
    status_p2h = P2H_STATUS_FOLLOW_UP if has_not_ok else P2H_STATUS_NORMAL
    follow_status = P2H_FOLLOW_NEW if has_not_ok else ""
    now = datetime.now().isoformat(timespec="seconds")
    upload = request.files.get("damage_photo")
    with db() as conn:
        insert_sql = """
            insert into p2h_reports (
                report_date, submit_time, driver_id, vehicle_id, odometer_start, fuel_status,
                general_note, damage_note, recommendation, status_p2h, follow_up_status,
                created_at, updated_at, created_by
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        if USE_POSTGRES:
            insert_sql += " returning id"
        cursor = conn.execute(
            insert_sql,
            (
                report_date,
                datetime.now().strftime("%H:%M"),
                driver["id"],
                int(vehicle_id),
                odometer_start,
                fuel_status,
                general_note,
                damage_note,
                recommendation,
                status_p2h,
                follow_status,
                now,
                now,
                emp["nik"],
            ),
        )
        report_id = cursor.fetchone()["id"] if USE_POSTGRES else cursor.lastrowid
        conn.executemany(
            "insert into p2h_checklist_items (p2h_report_id, category, item_name, result, note) values (?, ?, ?, ?, ?)",
            [(report_id, *item) for item in checklist_payload],
        )
        if upload and upload.filename:
            filename = secure_filename(upload.filename)
            ext = Path(filename).suffix.lower()
            if ext not in {".jpg", ".jpeg", ".png", ".webp", ".pdf"}:
                conn.rollback()
                return error_response("File foto harus JPG, PNG, WEBP, atau PDF.", "p2h-checklist", 400)
            P2H_ATTACHMENT_DIR.mkdir(parents=True, exist_ok=True)
            stored_name = f"p2h_{report_id}_{int(datetime.now().timestamp())}{ext}"
            upload.save(P2H_ATTACHMENT_DIR / stored_name)
            conn.execute(
                "insert into p2h_attachments (p2h_report_id, file_path, original_name, uploaded_at) values (?, ?, ?, ?)",
                (report_id, stored_name, filename, now),
            )
        conn.commit()
    return ok_response("p2h-checklist", "Checklist P2H berhasil dikirim")


@app.post("/p2h/<int:report_id>/follow-up")
def update_p2h_follow_up(report_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    report = row("select id from p2h_reports where id = ?", (report_id,))
    if not report:
        return error_response("Laporan P2H tidak ditemukan.", "p2h-report", 404)
    follow_up_status = request.form.get("follow_up_status", P2H_FOLLOW_PROCESS).strip()
    if follow_up_status not in {P2H_FOLLOW_NEW, P2H_FOLLOW_PROCESS, P2H_FOLLOW_DONE, P2H_FOLLOW_REJECTED}:
        follow_up_status = P2H_FOLLOW_PROCESS
    execute(
        """
        update p2h_reports
        set follow_up_status = ?, follow_up_note = ?, follow_up_action = ?, follow_up_date = ?, updated_at = ?
        where id = ?
        """,
        (
            follow_up_status,
            request.form.get("follow_up_note", "").strip(),
            request.form.get("follow_up_action", "").strip(),
            request.form.get("follow_up_date", "").strip() or today_iso(),
            datetime.now().isoformat(timespec="seconds"),
            report_id,
        ),
    )
    return ok_response("p2h-report", "Follow up P2H berhasil disimpan")


@app.post("/p2h/workdays")
def save_p2h_workdays():
    emp = current_employee()
    if not has_role(emp, "ga_admin", "super_admin"):
        return error_response("Unauthorized", "p2h-report", 403)
    data = request.get_json(silent=True) or request.form
    start_month = (data.get("start_month") or "").strip()
    end_month = (data.get("end_month") or "").strip()
    try:
        workdays = int(data.get("workdays") or 0)
        datetime.strptime(start_month + "-01", "%Y-%m-%d")
        datetime.strptime(end_month + "-01", "%Y-%m-%d")
    except (TypeError, ValueError):
        return error_response("Invalid request", "p2h-report", 400)
    if workdays < 0 or workdays > 366:
        return error_response("Invalid request", "p2h-report", 400)
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute(
            """
            insert into p2h_workday_overrides (start_month, end_month, workdays, updated_by, updated_at)
            values (?, ?, ?, ?, ?)
            on conflict(start_month, end_month)
            do update set workdays = excluded.workdays, updated_by = excluded.updated_by, updated_at = excluded.updated_at
            """,
            (start_month, end_month, workdays, emp["nik"], now),
        )
    return ok_response("p2h-report", "Hari kerja P2H berhasil disimpan")


@app.get("/p2h/<int:report_id>/attachment")
def download_p2h_attachment(report_id: int):
    emp = current_employee()
    report = p2h_report_detail(report_id)
    if not report or not report.get("attachment_path"):
        return redirect(url_for("index"))
    is_owner = has_role(emp, "driver") and report.get("driver_nik") == emp.get("nik")
    if not (is_owner or has_role(emp, "ga_admin")):
        return redirect(url_for("index"))
    return send_from_directory(
        P2H_ATTACHMENT_DIR,
        report["attachment_path"],
        as_attachment=True,
        download_name=report.get("attachment_name") or report["attachment_path"],
    )


@app.post("/backups/create")
def create_backup():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"ga_operations_backup_{timestamp}.json" if USE_POSTGRES else f"ga_operations_backup_{timestamp}.db"
        backup_path = BACKUP_DIR / backup_name
        if USE_POSTGRES:
            create_postgres_json_backup(backup_path)
        else:
            shutil.copy2(DB_PATH, backup_path)
        app.logger.info("Database backup created file=%s by=%s", backup_name, emp.get("nik"))
        return ok_response("backup-restore", f"Backup berhasil dibuat: {backup_name}")
    except Exception as exc:
        app.logger.exception("Database backup failed")
        return error_response(f"Backup gagal: {exc}", "backup-restore", 500)


@app.get("/backups/<path:filename>/download")
def download_backup(filename: str):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    path = valid_backup_path(filename)
    if not path:
        return error_response("File backup tidak ditemukan.", "backup-restore", 404)
    return send_from_directory(BACKUP_DIR, path.name, as_attachment=True, download_name=path.name)


@app.post("/backups/<path:filename>/restore")
def restore_backup(filename: str):
    emp = current_employee()
    if not has_role(emp, "super_admin"):
        return error_response("Hanya Super Admin yang boleh restore database.", "backup-restore", 403)
    path = valid_backup_path(filename)
    if not path:
        return error_response("File backup tidak ditemukan.", "backup-restore", 404)
    try:
        if USE_POSTGRES:
            if path.suffix.lower() != ".json":
                return error_response("Restore PostgreSQL hanya mendukung backup JSON.", "backup-restore", 400)
            safety_name = f"pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            create_postgres_json_backup(BACKUP_DIR / safety_name)
            restore_postgres_json_backup(path)
        else:
            safety_name = f"pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy2(DB_PATH, BACKUP_DIR / safety_name)
            shutil.copy2(path, DB_PATH)
        app.logger.warning("Database restored from file=%s by=%s safety=%s", path.name, emp.get("nik"), safety_name)
        seed_data()
        return ok_response("backup-restore", f"Restore berhasil dari backup: {path.name}")
    except Exception as exc:
        app.logger.exception("Database restore failed")
        return error_response(f"Restore gagal: {exc}", "backup-restore", 500)


@app.post("/backups/<path:filename>/delete")
def delete_backup(filename: str):
    emp = current_employee()
    if not has_role(emp, "super_admin"):
        return error_response("Hanya Super Admin yang boleh menghapus cadangan database.", "backup-restore", 403)
    path = valid_backup_path(filename)
    if not path:
        return error_response("File backup tidak ditemukan.", "backup-restore", 404)
    try:
        deleted_name = path.name
        path.unlink()
        app.logger.warning("Database backup deleted file=%s by=%s", deleted_name, emp.get("nik"))
        return ok_response("backup-restore", f"File cadangan berhasil dihapus: {deleted_name}")
    except Exception as exc:
        app.logger.exception("Database backup delete failed")
        return error_response(f"Backup gagal: {exc}", "backup-restore", 500)


@app.post("/vehicles/<int:vehicle_id>/delete")
def delete_vehicle(vehicle_id: int):
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    execute("delete from vehicles where id = ?", (vehicle_id,))
    return redirect(url_for("index", tab="vehicles"))


@app.post("/trips/delete-all")
def delete_all_trips():
    emp = current_employee()
    if not has_role(emp, "super_admin"):
        return error_response("Only Super Admin can delete all trip history.", "dashboard", 403)
    with db() as conn:
        conn.execute("delete from trip_edit_logs")
        conn.execute("delete from trip_requests")
        conn.commit()
    return ok_response("dashboard", "All trip history deleted")


@app.get("/dev/reset-requests")
def dev_reset_requests():
    if not app.debug:
        return jsonify({"success": False, "message": "Development reset is only available when Flask debug mode is enabled."}), 403
    if request.args.get("confirm") != "yes":
        return jsonify({"success": False, "message": "Confirmation required. Use /dev/reset-requests?confirm=yes"}), 400

    allowed_tables = [
        "bookings",
        "booking_requests",
        "approvals",
        "trips",
        "trip_requests",
        "trip_history",
        "ratings",
        "driver_assignments",
        "edit_logs",
        "trip_edit_logs",
    ]
    forbidden_tables = {
        "employees",
        "users",
        "drivers",
        "vehicles",
        "roles",
        "employee_roles",
        "departments",
        "positions",
    }

    with db() as conn:
        existing_tables = list_tables(conn)
        tables_to_clear = [
            table for table in allowed_tables
            if table in existing_tables and table not in forbidden_tables
        ]
        logger.warning("Development reset will clear transaction tables only: %s", tables_to_clear)
        for table in tables_to_clear:
            conn.execute(f"delete from {table}")
        conn.commit()

    return jsonify({
        "success": True,
        "message": "Only booking/approval/trip test data has been reset. Master employee, driver, vehicle, and role data were not deleted.",
        "cleared_tables": tables_to_clear,
    })


@app.get("/employees/template.xlsx")
def employee_template():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Template"
    ws.append(["NIK", "Full Name", "Position", "Department", "Supervisor", "Phone Number", "Roles"])
    ws.append(["10001", "EMPLOYEE NAME", "Manager GA", "HRGA", "1001", "08123456789", "user,driver"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="employee_template.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/employees/export.xlsx")
def export_employees():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(["NIK", "Full Name", "Position", "Department", "Supervisor", "Phone Number", "Roles"])
    for item in rows(
        """
        select e.*, group_concat(er.role, ',') as roles_text
        from employees e left join employee_roles er on er.nik = e.nik
        where e.active = 1
        group by e.nik order by e.full_name
        """
    ):
        ws.append([item["nik"], item["full_name"], item["position"], item["department"], item["supervisor_nik"] or "", item["phone"], item["roles_text"] or "user"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="employees.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/employees/import")
def import_employees():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    file = request.files.get("employee_file")
    if not file:
        return error_response("Please select an Excel file.", "employees")
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    required = ["NIK", "Full Name", "Position", "Department", "Supervisor", "Phone Number", "Roles"]
    if headers[: len(required)] != required:
        return error_response("Invalid employee template columns.", "employees")
    valid_roles = set(ROLE_LABELS)
    imported = 0
    with db() as conn:
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            nik, full_name, position, department, supervisor, phone, roles_text = [str(value or "").strip() for value in row_values[:7]]
            if not nik or not full_name or not position or not department or not phone:
                continue
            roles = [role.strip() for role in (roles_text or "user").split(",") if role.strip()]
            if any(role not in valid_roles for role in roles):
                conn.rollback()
                return error_response(f"Invalid role for NIK {nik}.", "employees")
            if conn.execute("select nik from employees where nik = ?", (nik,)).fetchone():
                conn.rollback()
                return error_response(f"NIK sudah terdaftar: {nik}", "employees")
            position = " ".join(position.split())
            department = " ".join(department.split())
            conn.execute("insert or ignore into option_lists (kind, value) values ('position', ?)", (position,))
            conn.execute("insert or ignore into option_lists (kind, value) values ('department', ?)", (department,))
            conn.execute(
                "insert into employees (nik, full_name, position, department, supervisor_nik, phone) values (?, ?, ?, ?, ?, ?)",
                (nik, full_name.upper(), position, department, supervisor or None, phone),
            )
            for role_name in roles or ["user"]:
                conn.execute("insert into employee_roles (nik, role) values (?, ?)", (nik, role_name))
            sync_driver_record(conn, nik, full_name.upper(), position, phone, roles)
            imported += 1
        conn.commit()
    return ok_response("employees", f"Imported {imported} employees")


def schedule_summary() -> dict:
    today = datetime.now().date().isoformat()
    month_prefix = today[:7]
    assigned_trips = trip_rows("r.status in (?, ?)", (STATUS_ASSIGNED, STATUS_ON_TRIP))
    queue_trips = trip_rows("r.status in (?, ?, ?)", (STATUS_PENDING, STATUS_APPROVED, STATUS_PROCESSING))
    monthly_source = trip_rows("r.status in (?, ?, ?, ?, ?)", (STATUS_PENDING, STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP))
    daily = [item for item in assigned_trips if (item.get("start_date") or item["travel_date"]) <= today <= (item.get("end_date") or item["travel_date"])]
    queue_daily = [item for item in queue_trips if (item.get("start_date") or item["travel_date"]) <= today <= (item.get("end_date") or item["travel_date"])]
    monthly = monthly_source
    assigned_driver_ids = {item["driver_id"] for item in daily if item.get("driver_id")}
    all_drivers = rows("select * from drivers where status = 'ACTIVE' order by driver_name")
    return {
        "available_drivers": [driver for driver in all_drivers if driver["id"] not in assigned_driver_ids],
        "assigned_drivers": [driver for driver in all_drivers if driver["id"] in assigned_driver_ids],
        "drivers": all_drivers,
        "vehicles": rows("select * from vehicles order by plate_number"),
        "daily": daily,
        "queue_daily": queue_daily,
        "queue_monthly": queue_trips,
        "monthly": monthly,
    }


def vehicle_expiry_alerts() -> list[dict]:
    today = datetime.now().date()
    warning_until = today + timedelta(days=30)
    alerts = []
    vehicles = rows(
        """
        select id, plate_number, vehicle_name, stnk_expiry_date, kir_expiry_date
        from vehicles
        order by plate_number
        """
    )
    for vehicle in vehicles:
        for field, document in (("stnk_expiry_date", "STNK"), ("kir_expiry_date", "KIR")):
            raw_date = (vehicle.get(field) or "").strip()
            if not raw_date:
                continue
            try:
                expiry_date = datetime.fromisoformat(raw_date).date()
            except ValueError:
                logger.warning("Invalid %s expiry date for vehicle %s: %s", document, vehicle.get("id"), raw_date)
                continue
            if expiry_date <= warning_until:
                days_left = (expiry_date - today).days
                alerts.append({
                    "vehicle_id": vehicle["id"],
                    "plate_number": vehicle["plate_number"],
                    "vehicle_name": vehicle["vehicle_name"],
                    "document": document,
                    "expiry_date": raw_date,
                    "days_left": days_left,
                    "expired": days_left < 0,
                    "tab": "vehicles",
                })
    return sorted(alerts, key=lambda item: (item["expiry_date"], item["plate_number"], item["document"]))


def add_months(year_month: str, months: int) -> str | None:
    if not year_month:
        return None
    try:
        year, month = [int(part) for part in year_month.split("-", 1)]
    except ValueError:
        return None
    month_index = (year * 12 + (month - 1)) + int(months or 0)
    return f"{month_index // 12:04d}-{month_index % 12 + 1:02d}"


def maintenance_alerts() -> list[dict]:
    today = datetime.now().date()
    warning_until = today + timedelta(days=30)
    alerts = []
    vehicles = rows(
        """
        select id, plate_number, vehicle_name, vehicle_type, current_km, last_maintenance_month, last_maintenance_date, last_service_date,
               maintenance_km_interval, maintenance_month_interval
        from vehicles
        order by plate_number
        """
    )
    for vehicle in vehicles:
        interval_km = int(vehicle.get("maintenance_km_interval") or 0)
        current_km = int(vehicle.get("current_km") or 0)
        if interval_km > 0:
            next_due_km = ((current_km // interval_km) + 1) * interval_km
            remaining_km = next_due_km - current_km
            if remaining_km <= 1000:
                alerts.append({
                    "type": "maintenance_km",
                    "vehicle_id": vehicle["id"],
                    "plate_number": vehicle["plate_number"],
                    "vehicle_name": vehicle["vehicle_name"],
                    "document": "KM Service",
                    "expiry_date": str(next_due_km),
                    "days_left": remaining_km,
                    "expired": remaining_km <= 0,
                    "tab": "maintenance",
                    "message": f"{vehicle['plate_number']} mendekati jadwal service KM {next_due_km}. Sisa {max(remaining_km, 0)} KM.",
                })
        raw_maintenance_date = vehicle.get("last_maintenance_date") or vehicle.get("last_service_date") or (f"{vehicle.get('last_maintenance_month')}-01" if vehicle.get("last_maintenance_month") else "")
        due_month = add_months(raw_maintenance_date[:7], int(vehicle.get("maintenance_month_interval") or 0))
        if due_month:
            due_day = raw_maintenance_date[8:10] if len(raw_maintenance_date) >= 10 else "01"
            due_date = datetime.strptime(f"{due_month}-{due_day}", "%Y-%m-%d").date()
            days_left = (due_date - today).days
            if due_date <= warning_until:
                alerts.append({
                    "type": "maintenance_month",
                    "vehicle_id": vehicle["id"],
                    "plate_number": vehicle["plate_number"],
                    "vehicle_name": vehicle["vehicle_name"],
                    "document": "Bulan Service",
                    "expiry_date": due_month,
                    "days_left": days_left,
                    "expired": days_left < 0,
                    "tab": "maintenance",
                    "message": f"{vehicle['plate_number']} mendekati jadwal service bulan {due_month}.",
                })
    return alerts


def maintenance_history_rows() -> list[dict]:
    return rows(
        """
        select h.*, v.plate_number, v.vehicle_name, v.vehicle_type, e.full_name as created_by_name
        from vehicle_maintenance_history h
        join vehicles v on v.id = h.vehicle_id
        left join employees e on e.nik = h.created_by
        order by h.created_at desc, h.id desc
        """
    )


def driver_sim_alerts() -> list[dict]:
    today = datetime.now().date()
    warning_until = today + timedelta(days=30)
    alerts = []
    drivers = rows("select id, driver_name, sim_expiry_date from drivers where status = 'ACTIVE' order by driver_name")
    for driver in drivers:
        raw_date = (driver.get("sim_expiry_date") or "").strip()
        if not raw_date:
            continue
        try:
            expiry_date = datetime.fromisoformat(raw_date).date()
        except ValueError:
            logger.warning("Invalid SIM expiry date for driver %s: %s", driver.get("id"), raw_date)
            continue
        if expiry_date <= warning_until:
            days_left = (expiry_date - today).days
            alerts.append({
                "type": "driver_sim",
                "driver_id": driver["id"],
                "title": "SIM Driver",
                "message": f"SIM {driver['driver_name']} akan habis masa berlaku.",
                "expiry_date": raw_date,
                "days_left": days_left,
                "expired": days_left < 0,
                "tab": "vehicles",
            })
    return alerts


def average(values: list[float]) -> float:
    return round(sum(values) / len(values), 2) if values else 0


def month_key(value: str | None) -> str:
    return (value or "")[:7]


def fuel_consumption_transactions() -> list[dict]:
    raw_items = rows(
        """
        select r.id, r.request_code, r.vehicle_id, r.km_start, r.km_end, r.fuel_liters, r.cost_fuel,
               r.fuel_type, r.vehicle_condition_notes, coalesce(r.end_date, r.travel_date) as fuel_date,
               r.destination, r.purpose, d.driver_name, v.plate_number, v.vehicle_name, v.vehicle_type
        from trip_requests r
        join vehicles v on v.id = r.vehicle_id
        left join drivers d on d.id = r.driver_id
        where r.status = ?
          and r.vehicle_id is not null
          and r.km_end is not null
          and coalesce(r.fuel_liters, 0) > 0
        order by r.vehicle_id, fuel_date, r.id
        """,
        (STATUS_COMPLETED,),
    )
    grouped: dict[int, list[dict]] = {}
    for item in raw_items:
        grouped.setdefault(item["vehicle_id"], []).append(item)

    transactions = []
    for vehicle_id, items in grouped.items():
        previous_odometer = None
        for item in items:
            odometer = int(item.get("km_end") or 0)
            liters = float(item.get("fuel_liters") or 0)
            distance = 0
            if previous_odometer is not None and odometer > previous_odometer:
                distance = odometer - previous_odometer
            elif item.get("km_start") is not None and odometer > int(item.get("km_start") or 0):
                distance = odometer - int(item.get("km_start") or 0)
            consumption = round(distance / liters, 2) if distance > 0 and liters > 0 else 0
            cost = int(item.get("cost_fuel") or 0)
            transactions.append({
                "trip_id": item["id"],
                "request_code": item["request_code"],
                "vehicle_id": vehicle_id,
                "plate_number": item["plate_number"],
                "vehicle_name": item["vehicle_name"],
                "vehicle_type": item["vehicle_type"],
                "fuel_date": item["fuel_date"],
                "odometer": odometer,
                "previous_odometer": previous_odometer,
                "distance_km": distance,
                "fuel_liters": round(liters, 2),
                "fuel_price": round(cost / liters, 2) if liters else 0,
                "cost_fuel": cost,
                "driver_name": item.get("driver_name") or "-",
                "destination": item.get("destination") or "-",
                "purpose": item.get("purpose") or "-",
                "fuel_type": item.get("fuel_type") or "-",
                "vehicle_condition_notes": item.get("vehicle_condition_notes") or "",
                "km_per_liter": consumption,
            })
            if odometer > 0:
                previous_odometer = odometer
    return transactions


def vehicle_health_dashboard() -> list[dict]:
    vehicles = rows("select * from vehicles order by plate_number")
    transactions = fuel_consumption_transactions()
    today = datetime.now().date()
    month_this = today.strftime("%Y-%m")
    cutoff_1mo = (today - timedelta(days=30)).isoformat()
    cutoff_3mo = (today - timedelta(days=90)).isoformat()
    cutoff_6mo = (today - timedelta(days=180)).isoformat()
    health = []
    for vehicle in vehicles:
        vehicle_transactions = [item for item in transactions if item["vehicle_id"] == vehicle["id"]]
        valid_consumption = [item["km_per_liter"] for item in vehicle_transactions if item["km_per_liter"] > 0]
        baseline = average(valid_consumption)
        month_transactions = [item for item in vehicle_transactions if month_key(item["fuel_date"]) == month_this]
        transactions_1mo = [item for item in vehicle_transactions if (item.get("fuel_date") or "") >= cutoff_1mo]
        transactions_3mo = [item for item in vehicle_transactions if (item.get("fuel_date") or "") >= cutoff_3mo]
        transactions_6mo = [item for item in vehicle_transactions if (item.get("fuel_date") or "") >= cutoff_6mo]
        month_consumption = average([item["km_per_liter"] for item in month_transactions if item["km_per_liter"] > 0])
        consumption_1mo = average([item["km_per_liter"] for item in transactions_1mo if item["km_per_liter"] > 0])
        consumption_3mo = average([item["km_per_liter"] for item in transactions_3mo if item["km_per_liter"] > 0])
        consumption_6mo = average([item["km_per_liter"] for item in transactions_6mo if item["km_per_liter"] > 0])
        comparison = consumption_1mo or month_consumption or baseline
        drop_percent = round(((baseline - comparison) / baseline) * 100, 2) if baseline and comparison else 0
        if not valid_consumption:
            status = "Belum ada data"
            score = 0
        elif comparison >= baseline:
            status = "Efisien"
            score = min(100, round((comparison / baseline) * 100))
        elif drop_percent >= 20:
            status = "Kritis"
            score = max(0, round((comparison / baseline) * 100))
        elif drop_percent >= 10:
            status = "Warning"
            score = max(0, round((comparison / baseline) * 100))
        else:
            status = "Normal"
            score = max(0, round((comparison / baseline) * 100)) if baseline else 0
        monthly_summary = []
        for month in sorted({month_key(item["fuel_date"]) for item in vehicle_transactions if item.get("fuel_date")}):
            month_items = [item for item in vehicle_transactions if month_key(item["fuel_date"]) == month]
            monthly_summary.append({
                "month": month,
                "km_per_liter": average([item["km_per_liter"] for item in month_items if item["km_per_liter"] > 0]),
                "fuel_liters": round(sum(item["fuel_liters"] for item in month_items), 2),
                "cost_fuel": sum(item["cost_fuel"] for item in month_items),
            })
        health.append({
            "vehicle_id": vehicle["id"],
            "plate_number": vehicle["plate_number"],
            "vehicle_name": vehicle["vehicle_name"],
            "vehicle_type": vehicle["vehicle_type"],
            "current_km": vehicle.get("current_km") or 0,
            "status": status,
            "status_class": status.lower().replace(" ", "-"),
            "score": score,
            "alert": drop_percent >= 10,
            "alert_message": "Konsumsi BBM kendaraan ini turun dibanding rata-rata historis. Perlu pengecekan." if drop_percent >= 10 else "",
            "baseline_km_per_liter": baseline,
            "current_month_km_per_liter": month_consumption,
            "km_per_liter_1mo": consumption_1mo,
            "km_per_liter_3mo": consumption_3mo,
            "km_per_liter_6mo": consumption_6mo,
            "drop_percent": drop_percent,
            "fuel_liters_month": round(sum(item["fuel_liters"] for item in month_transactions), 2),
            "fuel_cost_month": sum(item["cost_fuel"] for item in month_transactions),
            "fuel_liters_1mo": round(sum(item["fuel_liters"] for item in transactions_1mo), 2),
            "fuel_cost_1mo": sum(item["cost_fuel"] for item in transactions_1mo),
            "fuel_liters_3mo": round(sum(item["fuel_liters"] for item in transactions_3mo), 2),
            "fuel_cost_3mo": sum(item["cost_fuel"] for item in transactions_3mo),
            "fuel_liters_6mo": round(sum(item["fuel_liters"] for item in transactions_6mo), 2),
            "fuel_cost_6mo": sum(item["cost_fuel"] for item in transactions_6mo),
            "monthly_trend": monthly_summary[-12:],
            "transactions": list(reversed(vehicle_transactions[-10:])),
        })
    return health


def task_notifications(emp: dict) -> list[dict]:
    notifications = []
    roles = normalize_roles(emp.get("roles"))
    if "pimpinan" in roles and "super_admin" not in roles:
        leader_items = trip_rows("e.supervisor_nik = ? and r.status = ?", (emp["nik"], STATUS_PENDING))
        for item in leader_items:
            notifications.append({
                "type": "approval",
                "title": "Approval Pimpinan",
                "message": f"{item['request_code']} menunggu approval.",
                "tab": "approval",
            })
    if has_role(emp, "ga_admin"):
        ga_items = trip_rows("r.status in (?, ?)", (STATUS_APPROVED, STATUS_PROCESSING))
        for item in ga_items:
            notifications.append({
                "type": "assignment",
                "title": "Assignment GA",
                "message": f"{item['request_code']} menunggu assignment driver/kendaraan.",
                "tab": "ga",
            })
    if has_role(emp, "driver"):
        driver_items = trip_rows("d.nik = ? and r.status = ?", (emp["nik"], STATUS_ASSIGNED))
        for item in driver_items:
            notifications.append({
                "type": "driver_start",
                "title": "Perjalanan Driver",
                "message": f"{item['request_code']} siap dimulai.",
                "tab": "driver",
            })
    if has_role(emp, "user"):
        review_items = trip_rows("r.requester_nik = ? and r.status = ? and coalesce(r.rating, 0) = 0", (emp["nik"], STATUS_COMPLETED))
        for item in review_items:
            notifications.append({
                "type": "review",
                "title": "Review Perjalanan",
                "message": f"{item['request_code']} menunggu review.",
                "tab": "history",
            })
    return notifications


def get_pending_departure_alerts(user_role, user_id, lang: str = "id") -> list[dict]:
    roles = normalize_roles(user_role)
    if "super_admin" in roles:
        return []
    is_ga = bool(roles.intersection({"ga_admin", "admin", "super_admin"}))
    is_driver = "driver" in roles
    if not is_ga and not is_driver:
        return []

    now = datetime.now()
    reminder_until = now + timedelta(minutes=45)
    today = now.date().isoformat()
    items = trip_rows(
        """
        coalesce(r.start_date, r.travel_date) = ?
        and r.status in (?, ?, ?)
        and r.driver_id is not null
        """,
        (today, STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED),
    )
    alerts = []
    for item in items:
        if is_driver and not is_ga and item.get("driver_nik") != user_id:
            continue
        try:
            departure_at = parse_dt(item.get("start_date") or item["travel_date"], item["depart_time"])
        except Exception:
            continue
        if departure_at > reminder_until:
            continue
        alert = {
            "id": item["id"],
            "request_code": item["request_code"],
            "full_name": item["full_name"],
            "destination": item["destination"],
            "depart_time": item["depart_time"],
            "driver_name": item.get("driver_name") or "-",
            "driver_nik": item.get("driver_nik"),
            "plate_number": item.get("plate_number") or "-",
            "vehicle_name": item.get("vehicle_name") or "",
            "status": item["status"],
            "status_label": status_label(item["status"], lang),
            "minutes_to_departure": max(0, int((departure_at - now).total_seconds() // 60)),
            "minimum_km": minimum_vehicle_start_km(item["vehicle_id"], item["id"]) if item.get("vehicle_id") else 0,
        }
        alerts.append(alert)
    return alerts


def performance_scoreboard() -> dict:
    completed = trip_rows("r.status = ? and coalesce(r.rating, 0) > 0", (STATUS_COMPLETED,))
    total = len(completed)
    avg = round(sum(item["rating"] or 0 for item in completed) / total, 2) if total else 0

    def level(score: float) -> str:
        if score >= 4.5:
            return "Excellence"
        if score >= 3.5:
            return "Good"
        return "Poor"

    drivers = {}
    for item in completed:
        key = item.get("driver_id") or 0
        if key not in drivers:
            drivers[key] = {"driver_name": item.get("driver_name") or "-", "ratings": [], "completed": 0}
        drivers[key]["ratings"].append(item["rating"] or 0)
        drivers[key]["completed"] += 1
    driver_scores = []
    for data in drivers.values():
        driver_avg = round(sum(data["ratings"]) / len(data["ratings"]), 2) if data["ratings"] else 0
        driver_scores.append({
            "driver_name": data["driver_name"],
            "average_rating": driver_avg,
            "completed_trips": data["completed"],
            "level": level(driver_avg),
        })
    return {
        "overall_score": avg,
        "average_rating": avg,
        "total_completed_trips": total,
        "level": level(avg) if total else "-",
        "drivers": sorted(driver_scores, key=lambda item: item["average_rating"], reverse=True),
    }


def performance_history() -> list[dict]:
    today = datetime.now().date().replace(day=1)
    months = []
    for offset in range(23, -1, -1):
        month_index = today.year * 12 + today.month - 1 - offset
        months.append(f"{month_index // 12:04d}-{month_index % 12 + 1:02d}")
    completed = trip_rows("r.status = ? and coalesce(r.rating, 0) > 0", (STATUS_COMPLETED,))
    history = []
    for month in months:
        ratings = [
            item["rating"] or 0 for item in completed
            if ((item.get("end_date") or item["travel_date"]) or "")[:7] == month
        ]
        history.append({
            "month": month,
            "average_rating": round(sum(ratings) / len(ratings), 2) if ratings else 0,
            "completed_trips": len(ratings),
        })
    return history


def performance_daily_history() -> dict:
    completed = trip_rows("r.status = ? and coalesce(r.rating, 0) > 0", (STATUS_COMPLETED,))
    months = {item["month"] for item in performance_history()}
    result: dict[str, list[dict]] = {}
    for month in months:
        year, month_num = [int(part) for part in month.split("-")]
        days_in_month = (datetime(year + (1 if month_num == 12 else 0), 1 if month_num == 12 else month_num + 1, 1).date() - timedelta(days=1)).day
        rows_for_month = []
        for day in range(1, days_in_month + 1):
            date_key = f"{month}-{day:02d}"
            ratings = [
                item["rating"] or 0 for item in completed
                if (item.get("end_date") or item["travel_date"]) == date_key
            ]
            rows_for_month.append({
                "date": date_key,
                "average_rating": round(sum(ratings) / len(ratings), 2) if ratings else 0,
                "completed_trips": len(ratings),
            })
        result[month] = rows_for_month
    return result


def p2h_report_rows(where: str = "", params: tuple = ()) -> list[dict]:
    query = """
        select p.*, d.driver_name, d.nik as driver_nik,
               v.plate_number, v.vehicle_name, v.vehicle_type,
               a.file_path as attachment_path, a.original_name as attachment_name,
               (
                   select count(*)
                   from p2h_checklist_items ci
                   where ci.p2h_report_id = p.id and ci.result = 'Tidak OK'
               ) as not_ok_count
        from p2h_reports p
        join drivers d on d.id = p.driver_id
        join vehicles v on v.id = p.vehicle_id
        left join p2h_attachments a on a.p2h_report_id = p.id
    """
    if where:
        query += " where " + where
    query += " order by p.report_date desc, p.submit_time desc, p.id desc"
    return rows(query, params)


def p2h_report_detail(report_id: int) -> dict | None:
    report = row(
        """
        select p.*, d.driver_name, d.nik as driver_nik,
               v.plate_number, v.vehicle_name, v.vehicle_type,
               a.file_path as attachment_path, a.original_name as attachment_name
        from p2h_reports p
        join drivers d on d.id = p.driver_id
        join vehicles v on v.id = p.vehicle_id
        left join p2h_attachments a on a.p2h_report_id = p.id
        where p.id = ?
        """,
        (report_id,),
    )
    if not report:
        return None
    report["items"] = rows("select * from p2h_checklist_items where p2h_report_id = ? order by category, id", (report_id,))
    return report


def today_iso() -> str:
    return datetime.now().date().isoformat()


def get_today_p2h_status(driver_id: int, vehicle_id: int | None = None) -> list[dict]:
    if vehicle_id:
        return p2h_report_rows("p.driver_id = ? and p.vehicle_id = ? and p.report_date = ?", (driver_id, vehicle_id, today_iso()))
    return p2h_report_rows("p.driver_id = ? and p.report_date = ?", (driver_id, today_iso()))


def driver_default_vehicle(driver_id: int) -> dict | None:
    return row(
        """
        select v.*
        from drivers d join vehicles v on v.id = d.default_vehicle_id
        where d.id = ?
        """,
        (driver_id,),
    )


def assigned_vehicles_for_driver_today(driver_id: int) -> list[dict]:
    return rows(
        """
        select distinct v.*
        from trip_requests r join vehicles v on v.id = r.vehicle_id
        where r.driver_id = ?
          and coalesce(r.start_date, r.travel_date) <= ?
          and coalesce(r.end_date, r.travel_date) >= ?
          and r.status in (?, ?, ?)
        order by v.plate_number
        """,
        (driver_id, today_iso(), today_iso(), STATUS_ASSIGNED, STATUS_ON_TRIP, STATUS_COMPLETED),
    )


def p2h_driver_vehicle_options(driver_id: int) -> list[dict]:
    seen = set()
    vehicles = []
    default_vehicle = driver_default_vehicle(driver_id)
    if default_vehicle:
        vehicles.append(default_vehicle)
        seen.add(default_vehicle["id"])
    for vehicle in assigned_vehicles_for_driver_today(driver_id):
        if vehicle["id"] not in seen:
            vehicles.append(vehicle)
            seen.add(vehicle["id"])
    if not vehicles:
        vehicles = rows("select * from vehicles where status != 'MAINTENANCE' order by plate_number")
    return vehicles


def get_missing_p2h_drivers(date_value: str | None = None) -> list[dict]:
    date_value = date_value or today_iso()
    if datetime.now().time() < datetime.strptime("09:00", "%H:%M").time() and date_value == today_iso():
        return []
    active_drivers = rows("select * from drivers where status = 'ACTIVE' order by driver_name")
    submitted = {
        item["driver_id"] for item in rows("select distinct driver_id from p2h_reports where report_date = ?", (date_value,))
    }
    missing = []
    for driver in active_drivers:
        if driver["id"] in submitted:
            continue
        default_vehicle = driver_default_vehicle(driver["id"])
        missing.append({
            "driver_id": driver["id"],
            "driver_name": driver["driver_name"],
            "driver_nik": driver["nik"],
            "vehicle_id": default_vehicle["id"] if default_vehicle else None,
            "plate_number": default_vehicle["plate_number"] if default_vehicle else "-",
            "vehicle_name": default_vehicle["vehicle_name"] if default_vehicle else "-",
            "status": "Belum Submit",
            "deadline": "09:00",
        })
    return missing


def get_p2h_alerts(emp: dict) -> list[dict]:
    if has_role(emp, "super_admin"):
        return []
    missing = get_missing_p2h_drivers()
    if has_role(emp, "driver") and not has_role(emp, "ga_admin"):
        driver = row("select id from drivers where nik = ? and status = 'ACTIVE'", (emp["nik"],))
        if not driver:
            return []
        return [item for item in missing if item["driver_id"] == driver["id"]]
    if has_role(emp, "ga_admin"):
        return missing
    return []


def p2h_holidays(start_date: datetime.date | None = None, end_date: datetime.date | None = None) -> list[dict]:
    clauses = []
    params: list = []
    if start_date:
        clauses.append("holiday_date >= ?")
        params.append(start_date.isoformat())
    if end_date:
        clauses.append("holiday_date <= ?")
        params.append(end_date.isoformat())
    where = " where " + " and ".join(clauses) if clauses else ""
    return rows(f"select holiday_date, name from p2h_holidays{where} order by holiday_date", tuple(params))


def p2h_workday_overrides() -> list[dict]:
    return rows("select * from p2h_workday_overrides order by start_month desc, end_month desc")


def p2h_workday_override(start_month: str, end_month: str) -> int | None:
    item = row(
        "select workdays from p2h_workday_overrides where start_month = ? and end_month = ?",
        (start_month, end_month),
    )
    return int(item["workdays"]) if item else None


def business_days(start_date: datetime.date, end_date: datetime.date, holiday_dates: set[str] | None = None) -> int:
    if end_date < start_date:
        return 0
    holiday_dates = holiday_dates or set()
    total = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5 and current.isoformat() not in holiday_dates:
            total += 1
        current += timedelta(days=1)
    return total


def p2h_required_days(start_month: str, end_month: str, start: datetime.date, end: datetime.date) -> int:
    override = p2h_workday_override(start_month, end_month)
    if override is not None:
        return override
    capped_end = min(end, datetime.now().date())
    holidays = {item["holiday_date"] for item in p2h_holidays(start, capped_end)}
    return business_days(start, capped_end, holidays)


def parse_month_range(args) -> tuple[str, str, datetime.date, datetime.date]:
    now = datetime.now().date()
    start_month = args.get("start") or args.get("start_month") or f"{now.year:04d}-{now.month:02d}"
    end_month = args.get("end") or args.get("end_month") or start_month
    if args.get("start_year") and args.get("start_month"):
        start_month = f"{int(args.get('start_year')):04d}-{int(args.get('start_month')):02d}"
    if args.get("end_year") and args.get("end_month"):
        end_month = f"{int(args.get('end_year')):04d}-{int(args.get('end_month')):02d}"
    start = datetime.strptime(start_month + "-01", "%Y-%m-%d").date()
    end_year, end_mon = [int(part) for part in end_month.split("-")]
    end = (datetime(end_year + (1 if end_mon == 12 else 0), 1 if end_mon == 12 else end_mon + 1, 1).date() - timedelta(days=1))
    return start_month, end_month, start, end


def p2h_performance_data(filters=None) -> dict:
    filters = filters or {}
    start_month, end_month, start, end = parse_month_range(filters)
    driver_id = filters.get("driver_id") or ""
    vehicle_id = filters.get("vehicle_id") or ""
    status_p2h = filters.get("status_p2h") or ""
    follow_up_status = filters.get("follow_up_status") or ""
    clauses = ["p.report_date between ? and ?"]
    params: list = [start.isoformat(), end.isoformat()]
    if driver_id:
        clauses.append("p.driver_id = ?")
        params.append(driver_id)
    if vehicle_id:
        clauses.append("p.vehicle_id = ?")
        params.append(vehicle_id)
    if status_p2h:
        clauses.append("p.status_p2h = ?")
        params.append(status_p2h)
    if follow_up_status:
        clauses.append("coalesce(p.follow_up_status, '') = ?")
        params.append(follow_up_status)
    reports = p2h_report_rows(" and ".join(clauses), tuple(params))
    report_ids = [item["id"] for item in reports]
    not_ok_rows = []
    if report_ids:
        not_ok_rows = rows(
            f"select ci.*, p.report_date, p.driver_id, p.vehicle_id from p2h_checklist_items ci join p2h_reports p on p.id = ci.p2h_report_id where ci.result = 'Tidak OK' and ci.p2h_report_id in ({','.join('?' for _ in report_ids)})",
            tuple(report_ids),
        )
    drivers = rows("select * from drivers where status = 'ACTIVE' order by driver_name")
    if driver_id:
        drivers = [driver for driver in drivers if str(driver["id"]) == str(driver_id)]
    required_days = p2h_required_days(start_month, end_month, start, end)
    report_dates_by_driver: dict[int, set[str]] = {driver["id"]: set() for driver in drivers}
    not_ok_by_driver: dict[int, int] = {driver["id"]: 0 for driver in drivers}
    for report in reports:
        report_dates_by_driver.setdefault(report["driver_id"], set()).add(report["report_date"])
    for item in not_ok_rows:
        not_ok_by_driver[item["driver_id"]] = not_ok_by_driver.get(item["driver_id"], 0) + 1
    def consistency_label(value: float) -> str:
        if value >= 95:
            return "Sangat Konsisten"
        if value >= 85:
            return "Konsisten"
        if value >= 70:
            return "Perlu Monitoring"
        return "Tidak Konsisten"
    driver_performance = []
    for driver in drivers:
        submit_days = len(report_dates_by_driver.get(driver["id"], set()))
        consistency = round((submit_days / required_days * 100), 1) if required_days else 0
        default_vehicle = driver_default_vehicle(driver["id"])
        driver_performance.append({
            "driver_id": driver["id"],
            "driver_name": driver["driver_name"],
            "default_vehicle": f"{default_vehicle['plate_number']} - {default_vehicle['vehicle_name']}" if default_vehicle else "-",
            "required_days": required_days,
            "submitted_days": submit_days,
            "missing_days": max(required_days - submit_days, 0),
            "consistency": consistency,
            "not_ok_count": not_ok_by_driver.get(driver["id"], 0),
            "status": consistency_label(consistency),
        })
    months = []
    index = start.year * 12 + start.month - 1
    end_index = end.year * 12 + end.month - 1
    while index <= end_index:
        months.append(f"{index // 12:04d}-{index % 12 + 1:02d}")
        index += 1
    trend = []
    for month in months:
        month_reports = [item for item in reports if item["report_date"][:7] == month]
        trend.append({
            "month": month,
            "total": len(month_reports),
            "normal": len([item for item in month_reports if item["status_p2h"] == P2H_STATUS_NORMAL]),
            "follow_up": len([item for item in month_reports if item["status_p2h"] == P2H_STATUS_FOLLOW_UP]),
            "not_ok": len([item for item in not_ok_rows if item["report_date"][:7] == month]),
        })
    vehicle_findings = {}
    for item in not_ok_rows:
        report = next((report for report in reports if report["id"] == item["p2h_report_id"]), None)
        if report:
            key = report["plate_number"]
            vehicle_findings[key] = vehicle_findings.get(key, 0) + 1
    summary = {
        "required_days": required_days,
        "total_reports": len(reports),
        "average_consistency": round(sum(item["consistency"] for item in driver_performance) / len(driver_performance), 1) if driver_performance else 0,
        "missing_drivers_today": len(get_missing_p2h_drivers()) if start <= datetime.now().date() <= end else 0,
        "total_not_ok": len(not_ok_rows),
        "follow_up_reports": len([item for item in reports if item["status_p2h"] == P2H_STATUS_FOLLOW_UP]),
        "follow_up_done": len([item for item in reports if item["follow_up_status"] == P2H_FOLLOW_DONE]),
        "follow_up_pending": len([item for item in reports if item["status_p2h"] == P2H_STATUS_FOLLOW_UP and item["follow_up_status"] != P2H_FOLLOW_DONE]),
    }
    return {
        "filters": {"start": start_month, "end": end_month},
        "summary": summary,
        "trend": trend,
        "driver_performance": sorted(driver_performance, key=lambda item: item["consistency"], reverse=True),
        "vehicle_findings": sorted([{"vehicle": key, "count": value} for key, value in vehicle_findings.items()], key=lambda item: item["count"], reverse=True),
    }


GUIDE_PDF_CONTENT = {
    "id": [
        ("Scope Role User", [
            "Membuat booking kendaraan, melihat riwayat pribadi, dan memberi review setelah perjalanan selesai.",
            "Melihat Dasbor Jadwal Driver sebagai referensi ketersediaan waktu.",
            "Booking baru dapat dibatasi jika masih ada perjalanan aktif atau selesai yang belum direview.",
        ]),
        ("Scope Role Pimpinan", [
            "Menerima request sesuai routing pimpinan user.",
            "Menyetujui atau menolak request dan mengisi catatan persetujuan.",
            "Super Admin tetap punya akses penuh, namun notifikasi approval normal hanya ke pimpinan terkait.",
        ]),
        ("Scope Role GA Admin", [
            "Mengelola request yang disetujui, memilih driver dan kendaraan, serta memastikan tidak ada konflik jadwal.",
            "Mengelola master kendaraan, driver, perawatan berkala, kesehatan kendaraan, review user, P2H, dan backup.",
            "Kendaraan maintenance tidak boleh dipakai. Kendaraan assigned/dedicated boleh dipilih dengan koordinasi pengguna terkait.",
        ]),
        ("Scope Role Driver", [
            "Melihat tugas perjalanan yang ditugaskan.",
            "Menekan Mulai Perjalanan, lalu menyelesaikan perjalanan dengan KM akhir, liter BBM, biaya BBM, tol, dan parkir.",
            "Mengisi Checklist P2H harian dan melaporkan kondisi tidak normal ke GA.",
        ]),
        ("Flow Booking Kendaraan", [
            "User mengisi tujuan, keperluan, tanggal, jam, jumlah penumpang, checklist plat, Google Maps, dan catatan.",
            "Sistem mengecek jadwal bentrok, kapasitas kursi, plat ganjil/genap/bebas, driver, kendaraan, dan buffer operasional.",
            "Jika penuh, user diarahkan melihat Dasbor Jadwal Driver dan menghubungi GA untuk kebutuhan mendesak.",
        ]),
        ("Flow Approval", [
            "Request baru masuk ke status Menunggu Persetujuan Pimpinan.",
            "Pimpinan dapat setuju atau tolak dari menu Persetujuan.",
            "Setelah disetujui, request masuk ke Kontrol GA untuk assignment driver dan kendaraan.",
        ]),
        ("Flow Kontrol GA", [
            "Sistem memberi rekomendasi kendaraan sesuai kapasitas, plat, jadwal, dan ketersediaan.",
            "Kendaraan default driver direkomendasikan bila tersedia, namun GA tetap bisa mengganti manual.",
            "Sistem menolak double booking driver atau kendaraan pada jadwal yang bentrok.",
        ]),
        ("Review, Perawatan, P2H, dan Backup", [
            "User wajib memberi rating dan komentar setelah perjalanan selesai.",
            "GA dapat mengisi tindak lanjut review, upload bukti PDF, dan export Excel.",
            "Perawatan berkala memonitor KM, tanggal service, referensi maintenance, dan link referensi.",
            "P2H wajib diisi driver setiap hari operasional; laporan tidak normal masuk ke GA untuk follow up.",
            "Backup database dapat dibuat GA Admin/Super Admin, sedangkan restore hanya Super Admin.",
        ]),
        ("Perhitungan Konsistensi P2H", [
            "Hari Wajib P2H dihitung dari hari kerja Senin-Jumat pada interval monitoring yang dipilih.",
            "Tanggal merah nasional/cuti bersama yang terdaftar di sistem tidak dihitung sebagai Hari Wajib P2H.",
            "Jika periode melewati masa depan, hitungan hari wajib hanya sampai tanggal hari ini.",
            "Admin GA dan Super Admin dapat mengisi manual jumlah hari kerja bila kalender operasional perusahaan berbeda.",
            "Jumlah Kirim P2H dihitung satu kali per driver per tanggal.",
            "Rumus: Jumlah Kirim P2H / Hari Wajib P2H x 100%. Contoh 2 dari 17 hari wajib = 11,8%.",
            "Kategori: 95-100% Sangat Konsisten, 85-94% Konsisten, 70-84% Perlu Monitoring, di bawah 70% Tidak Konsisten.",
        ]),
    ],
    "en": [
        ("User Role Scope", [
            "Create vehicle bookings, view personal history, and submit reviews after completed trips.",
            "Open Driver Schedule Dashboard to check available time slots.",
            "New bookings may be limited when active or unreviewed completed trips still exist.",
        ]),
        ("Leader Role Scope", [
            "Receive requests based on requester leader routing.",
            "Approve or reject requests and add approval notes.",
            "Super Admin has full access, but normal approval notifications go only to the assigned leader.",
        ]),
        ("GA Admin Role Scope", [
            "Manage approved requests, assign drivers and vehicles, and prevent schedule conflicts.",
            "Manage vehicle master data, drivers, preventive maintenance, vehicle health, user reviews, P2H, and backups.",
            "Maintenance vehicles cannot be used. Assigned/dedicated vehicles may be selected with coordination to the related user.",
        ]),
        ("Driver Role Scope", [
            "View assigned trip tasks.",
            "Start trips, then complete trips with end KM, fuel liters, fuel cost, toll, and parking cost.",
            "Submit daily P2H Checklist and report abnormal vehicle conditions to GA.",
        ]),
        ("Vehicle Booking Flow", [
            "User fills destination, purpose, dates, times, passenger count, plate rule, Google Maps, and notes.",
            "System checks conflicts, seat capacity, odd/even/free plate rule, driver, vehicle, and operational buffer.",
            "If full, user is directed to Driver Schedule Dashboard and GA contact for urgent needs.",
        ]),
        ("Approval Flow", [
            "New requests enter Pending Leader Approval status.",
            "Leader can approve or reject from the Approval menu.",
            "After approval, the request moves to GA Control for driver and vehicle assignment.",
        ]),
        ("GA Control Flow", [
            "System recommends vehicles based on capacity, plate rule, schedule, and availability.",
            "Driver default vehicle is recommended when available, but GA may change it manually.",
            "System blocks double booking for drivers or vehicles on overlapping schedules.",
        ]),
        ("Reviews, Maintenance, P2H, and Backup", [
            "Users must submit ratings and comments after completed trips.",
            "GA can enter review follow-up notes, upload PDF proof, and export Excel.",
            "Preventive maintenance monitors KM, service date, maintenance reference, and reference links.",
            "Drivers must submit P2H every operational day; abnormal reports are followed up by GA.",
            "GA Admin/Super Admin can create backups, while restore is Super Admin-only.",
        ]),
        ("P2H Consistency Calculation", [
            "Required P2H Days are counted from weekdays Monday-Friday in the selected monitoring interval.",
            "Registered national holidays/joint leave days are excluded from Required P2H Days.",
            "If the period extends into the future, required days are capped at today's date.",
            "GA Admin and Super Admin can manually override working days when company operations differ from the calendar.",
            "Submitted P2H Days are counted once per driver per date.",
            "Formula: Submitted P2H Days / Required P2H Days x 100%. Example: 2 of 17 required days = 11.8%.",
            "Categories: 95-100% Very Consistent, 85-94% Consistent, 70-84% Needs Monitoring, below 70% Not Consistent.",
        ]),
    ],
}


def pdf_escape(text: str) -> str:
    return str(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_wrap(text: str, max_chars: int) -> list[str]:
    words = str(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) > max_chars and current:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines or [""]


def guide_pdf_bytes(language: str) -> io.BytesIO:
    language = "en" if language == "en" else "id"
    title = "Guide & Technical - GA Operations System" if language == "en" else "Panduan & Teknis - GA Operations System"
    subtitle = "Compact user guide by role and workflow" if language == "en" else "Panduan ringkas berdasarkan role dan alur kerja"
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    page_width, page_height = 595, 842
    margin = 42
    pages: list[list[str]] = []
    commands: list[str] = []
    y = page_height - margin

    def add_page():
        nonlocal commands, y
        if commands:
            pages.append(commands)
        commands = []
        y = page_height - margin

    def add_line(text: str, size: int = 10, font: str = "F1", leading: int = 14, x: int = margin):
        nonlocal y
        if y < margin + leading:
            add_page()
        commands.append(f"BT /{font} {size} Tf {x} {y} Td ({pdf_escape(text)}) Tj ET")
        y -= leading

    add_line(title, 18, "F2", 24)
    add_line(subtitle, 10, "F1", 16)
    add_line(f"Generated: {generated}", 9, "F1", 18)
    for section_title, points in GUIDE_PDF_CONTENT[language]:
        if y < 120:
            add_page()
        add_line(section_title, 12, "F2", 17)
        for point in points:
            wrapped = pdf_wrap(point, 92)
            add_line(f"- {wrapped[0]}", 9, "F1", 12, margin + 8)
            for continuation in wrapped[1:]:
                add_line(f"  {continuation}", 9, "F1", 12, margin + 8)
        y -= 4
    if commands:
        pages.append(commands)

    objects: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    ]
    page_refs = []
    for page_commands in pages:
        page_obj_num = len(objects) + 1
        content_obj_num = page_obj_num + 1
        page_refs.append(f"{page_obj_num} 0 R")
        content = "\n".join(page_commands).encode("cp1252", errors="replace")
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_obj_num} 0 R >>".encode())
        objects.append(b"<< /Length " + str(len(content)).encode() + b" >>\nstream\n" + content + b"\nendstream")
    objects[1] = f"<< /Type /Pages /Kids [{' '.join(page_refs)}] /Count {len(page_refs)} >>".encode()

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode())
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref_at = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode())
    pdf.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_at}\n%%EOF".encode())
    buffer = io.BytesIO(bytes(pdf))
    buffer.seek(0)
    return buffer


@app.get("/api/data")
def api_data():
    emp = current_employee()
    if not emp:
        return jsonify({"authenticated": False})

    data: dict = {
        "authenticated": True,
        "me": emp,
        "language": current_language(),
        "roles": emp["roles"],
        "role_labels": ROLE_LABELS,
        "status_labels": {status: status_label(status) for status in [
            STATUS_PENDING, STATUS_REJECTED, STATUS_APPROVED, STATUS_PROCESSING,
            STATUS_ASSIGNED, STATUS_ON_TRIP, STATUS_COMPLETED, STATUS_REVIEWED, STATUS_CANCELED
        ]},
        "requests": [],
        "dashboard_requests": [],
        "employees": [],
        "drivers": [],
        "vehicles": [],
        "vehicle_alerts": [],
        "driver_sim_alerts": [],
        "maintenance_alerts": [],
        "departure_alerts": [],
        "notifications": [],
        "maintenance_vehicles": [],
        "maintenance_history": [],
        "maintenance_type_options": MAINTENANCE_TYPE_OPTIONS,
        "maintenance_part_options": MAINTENANCE_PART_OPTIONS,
        "user_reviews": [],
        "p2h_checklist": P2H_CHECKLIST,
        "p2h_reports": [],
        "p2h_my_reports": [],
        "p2h_vehicle_options": [],
        "p2h_alerts": [],
        "p2h_performance": {},
        "p2h_holidays": [],
        "p2h_workday_overrides": [],
        "backup_history": [],
        "can_restore_backup": has_role(emp, "super_admin"),
        "can_edit_vehicle_km": has_role(emp, "super_admin"),
        "can_delete_maintenance_history": has_role(emp, "super_admin"),
        "schedule": schedule_summary(),
        "scoreboard": performance_scoreboard(),
        "performance_history": performance_history(),
        "performance_daily_history": performance_daily_history(),
        "vehicle_health": [],
        "options": {
            "positions": option_values("position"),
            "departments": option_values("department"),
        },
        "stats": {},
    }
    data["departure_alerts"] = get_pending_departure_alerts(emp["roles"], emp["nik"], data["language"])

    if has_role(emp, "user"):
        data["requests"] = trip_rows("r.requester_nik = ?", (emp["nik"],))
    if has_role(emp, "driver"):
        driver = row("select * from drivers where nik = ? and status = 'ACTIVE'", (emp["nik"],))
        if driver:
            data["p2h_vehicle_options"] = p2h_driver_vehicle_options(driver["id"])
            data["p2h_my_reports"] = p2h_report_rows("p.driver_id = ?", (driver["id"],))
    if has_role(emp, "ga_admin", "pimpinan"):
        data["dashboard_requests"] = trip_rows()
    if has_role(emp, "pimpinan"):
        leader_items = trip_rows("(e.supervisor_nik = ? or ? = 1)", (emp["nik"], 1 if has_role(emp, "super_admin") else 0))
        data["leader_requests"] = leader_items
    if has_role(emp, "ga_admin"):
        data["ga_requests"] = trip_rows("r.status in (?, ?, ?, ?)", (STATUS_APPROVED, STATUS_PROCESSING, STATUS_ASSIGNED, STATUS_ON_TRIP))
        for item in data["ga_requests"]:
            item["availability"] = availability_for_trip(item)
        data["employees"] = rows(
            """
            select e.*, group_concat(er.role, ', ') as roles_text
            from employees e left join employee_roles er on er.nik = e.nik
            where e.active = 1
            group by e.nik order by e.full_name
            """
        )
        data["drivers"] = rows("select * from drivers where status = 'ACTIVE' order by driver_name")
        data["vehicles"] = rows("select * from vehicles order by plate_number")
        data["maintenance_vehicles"] = rows("select * from vehicles order by plate_number")
        data["maintenance_history"] = maintenance_history_rows()
        data["user_reviews"] = trip_rows("coalesce(r.rating, 0) > 0")
        data["vehicle_alerts"] = vehicle_expiry_alerts()
        data["driver_sim_alerts"] = driver_sim_alerts()
        data["maintenance_alerts"] = maintenance_alerts()
        data["vehicle_health"] = vehicle_health_dashboard()
        data["p2h_reports"] = p2h_report_rows()
        data["p2h_performance"] = p2h_performance_data(request.args)
        data["p2h_holidays"] = p2h_holidays()
        data["p2h_workday_overrides"] = p2h_workday_overrides()
        data["backup_history"] = backup_history()
    data["p2h_alerts"] = get_p2h_alerts(emp)
    data["notifications"] = task_notifications(emp)
    if has_role(emp, "super_admin"):
        data["notifications"] = []
        data["departure_alerts"] = []
        data["p2h_alerts"] = []
    else:
        for item in data["departure_alerts"]:
            is_driver_alert = has_role(emp, "driver") and item.get("driver_nik") == emp["nik"] and not has_role(emp, "ga_admin")
            data["notifications"].append({
                "type": "departure_reminder",
                "title": "Reminder Keberangkatan",
                "message": (
                    f"Jadwal pukul {item['depart_time']} belum dimulai."
                    if is_driver_alert
                    else f"{item['request_code']} belum dimulai pukul {item['depart_time']}."
                ),
                "tab": "driver" if is_driver_alert else "dashboard",
            })
    if has_role(emp, "ga_admin") and not has_role(emp, "super_admin"):
        data["notifications"].extend(data["driver_sim_alerts"])
        data["notifications"].extend(data["vehicle_alerts"])
        data["notifications"].extend(data["maintenance_alerts"])
        for item in data["vehicle_health"]:
            if item.get("alert"):
                data["notifications"].append({
                    "type": "fuel_efficiency",
                    "title": "Konsumsi BBM",
                    "message": f"{item['plate_number']} boros dibanding rata-rata historis.",
                    "expiry_date": f"-{item['drop_percent']}%",
                    "tab": "dashboard",
                })
    if data["p2h_alerts"]:
        if has_role(emp, "driver") and not has_role(emp, "ga_admin"):
            data["notifications"].append({
                "type": "p2h_missing",
                "title": "Checklist P2H",
                "message": "Anda belum mengisi Checklist P2H hari ini.",
                "tab": "p2h-checklist",
            })
        elif has_role(emp, "ga_admin") and not has_role(emp, "super_admin"):
            data["notifications"].append({
                "type": "p2h_missing_ga",
                "title": "Checklist P2H",
                "message": "Terdapat driver yang belum mengisi P2H hari ini.",
                "tab": "p2h-report",
            })
    if has_role(emp, "driver"):
        data["driver_requests"] = trip_rows("d.nik = ? and r.status in (?, ?, ?)", (emp["nik"], STATUS_ASSIGNED, STATUS_ON_TRIP, STATUS_COMPLETED))

    all_trips = trip_rows() if has_role(emp, "ga_admin", "pimpinan") else data["requests"]
    data["stats"] = {
        "total": len(all_trips),
        "pending": sum(1 for x in all_trips if x["status"] == STATUS_PENDING),
        "assigned": sum(1 for x in all_trips if x["status"] in [STATUS_ASSIGNED, STATUS_ON_TRIP]),
        "completed": sum(1 for x in all_trips if x["status"] == STATUS_COMPLETED),
        "review_required": row(
            "select count(*) as total from trip_requests where requester_nik = ? and status = ? and coalesce(rating, 0) = 0",
            (emp["nik"], STATUS_COMPLETED),
        )["total"],
    }
    return jsonify(data)


@app.get("/api/vehicle-health")
def api_vehicle_health():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return jsonify({"success": False, "message": "Unauthorized"}), 403
    return jsonify({
        "success": True,
        "items": vehicle_health_dashboard(),
        "transactions": fuel_consumption_transactions(),
    })


@app.get("/api/p2h/performance")
@app.get("/api/p2h/trend")
@app.get("/api/p2h/driver-consistency")
@app.get("/api/p2h/finding-summary")
def api_p2h_performance():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return jsonify({"success": False, "message": "Unauthorized"}), 403
    return jsonify({"success": True, **p2h_performance_data(request.args)})


@app.get("/guide/export.pdf")
def export_guide_pdf():
    emp = current_employee()
    if not emp:
        return redirect(url_for("login"))
    language = current_language()
    filename = "guide_technical_ga_operations.pdf" if language == "en" else "panduan_teknis_ga_operations.pdf"
    return send_file(
        guide_pdf_bytes(language),
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@app.get("/export/trips.xlsx")
def export_trips():
    emp = current_employee()
    if not has_role(emp, "ga_admin", "pimpinan"):
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Trip History"
    headers = [
        "Request Code", "Requester", "Department", "Destination", "Purpose", "Start Date", "End Date", "Depart",
        "Return", "Passengers", "Status", "Driver", "Vehicle", "KM Start", "KM End", "Fuel (Liter)", "Fuel (Rp)",
        "Toll", "Rating", "Review",
    ]
    ws.append(headers)
    for item in trip_rows():
        ws.append([
            item["request_code"], item["full_name"], item["department"], item["destination"], item["purpose"],
            item.get("start_date") or item["travel_date"], item.get("end_date") or item["travel_date"], item["depart_time"], item["return_time"], item["passengers"], item["status"],
            item["driver_name"] or "", item["plate_number"] or "", item["km_start"] or "", item["km_end"] or "",
            item.get("fuel_liters") or 0, item["cost_fuel"] or 0, item["cost_toll"] or 0, item["rating"] or "", item["review"] or "",
        ])
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 34)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="ga_trip_history.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/export/reviews.xlsx")
def export_reviews():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    if start and end and start > end:
        start, end = end, start
    wb = Workbook()
    ws = wb.active
    ws.title = "User Reviews"
    headers = [
        "Review Date", "Request Code", "Requester", "Department", "Destination", "Purpose",
        "Start Date", "End Date", "Driver", "Vehicle", "Rating", "Review", "Follow Up", "Proof File", "Proof Uploaded At",
    ]
    ws.append(headers)
    for item in trip_rows("coalesce(r.rating, 0) > 0"):
        review_month = ((item.get("updated_at") or item.get("end_date") or item.get("travel_date") or "")[:7])
        if start and review_month < start:
            continue
        if end and review_month > end:
            continue
        ws.append([
            (item.get("updated_at") or "")[:10],
            item["request_code"],
            item["full_name"],
            item["department"],
            item["destination"],
            item["purpose"],
            item.get("start_date") or item["travel_date"],
            item.get("end_date") or item["travel_date"],
            item.get("driver_name") or "",
            item.get("plate_number") or "",
            item.get("rating") or "",
            item.get("review") or "",
            item.get("review_follow_up") or "",
            item.get("review_proof_original_name") or "",
            item.get("review_proof_uploaded_at") or "",
        ])
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="ga_user_reviews.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/export/maintenance.xlsx")
def export_maintenance():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance History"
    headers = [
        "Tanggal Service", "Tanggal Referensi Maintenance", "Nopol", "Merk Kendaraan", "Tipe Kendaraan",
        "KM Service", "Jenis Perawatan", "Checklist Part", "Custom Jenis Perawatan", "Custom Part",
        "Input Oleh", "Created At",
    ]
    ws.append(headers)
    for item in maintenance_history_rows():
        try:
            maintenance_types = ", ".join(json.loads(item.get("maintenance_types") or "[]"))
        except json.JSONDecodeError:
            maintenance_types = item.get("maintenance_types") or ""
        try:
            parts = ", ".join(json.loads(item.get("parts") or "[]"))
        except json.JSONDecodeError:
            parts = item.get("parts") or ""
        ws.append([
            item.get("service_date") or "",
            item.get("maintenance_reference_date") or "",
            item.get("plate_number") or "",
            item.get("vehicle_name") or "",
            item.get("vehicle_type") or "",
            item.get("km_at_service") or 0,
            maintenance_types,
            parts,
            item.get("custom_maintenance_type") or "",
            item.get("custom_part") or "",
            item.get("created_by_name") or item.get("created_by") or "",
            item.get("created_at") or "",
        ])
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="maintenance_history.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/export/p2h.xlsx")
def export_p2h():
    emp = current_employee()
    if not has_role(emp, "ga_admin"):
        return redirect(url_for("index"))
    start_month, end_month, start, end = parse_month_range(request.args)
    clauses = ["p.report_date between ? and ?"]
    params: list = [start.isoformat(), end.isoformat()]
    for field in ["driver_id", "vehicle_id", "status_p2h", "follow_up_status"]:
        value = request.args.get(field, "").strip()
        if value:
            column = "coalesce(p.follow_up_status, '')" if field == "follow_up_status" else f"p.{field}"
            clauses.append(f"{column} = ?")
            params.append(value)
    reports = p2h_report_rows(" and ".join(clauses), tuple(params))
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan P2H"
    checklist_names = [item for items in P2H_CHECKLIST.values() for item in items]
    headers = [
        "Tanggal", "Jam Submit", "Driver", "Nopol", "Merk", "Tipe", "KM Awal",
        "Status P2H", "Jumlah Tidak OK", "Catatan Kerusakan", "Rekomendasi",
        "Status Follow Up", "Catatan Follow Up", "Tindakan Follow Up",
        *checklist_names,
    ]
    ws.append(headers)
    for report in reports:
        items = {
            item["item_name"]: item["result"]
            for item in rows("select item_name, result from p2h_checklist_items where p2h_report_id = ?", (report["id"],))
        }
        ws.append([
            report["report_date"],
            report["submit_time"],
            report["driver_name"],
            report["plate_number"],
            report["vehicle_name"],
            report["vehicle_type"],
            report["odometer_start"],
            report["status_p2h"],
            report["not_ok_count"],
            report.get("damage_note") or "",
            report.get("recommendation") or "",
            report.get("follow_up_status") or "",
            report.get("follow_up_note") or "",
            report.get("follow_up_action") or "",
            *[items.get(name, "") for name in checklist_names],
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="laporan_p2h.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    seed_data()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
